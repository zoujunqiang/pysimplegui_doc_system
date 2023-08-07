# encoding:utf-8
"""
@file = main
@author = zouju
@create_time = 2023-08-07- 11:06
"""


import time
from openpyxl import Workbook
from texttable import Texttable
from API.CommonApi.common_api import computerMD5, write_row_by_list, myPopen, copyfile, make_qr_code, get_FileSize, \
    upload_to_server, write_db_from_file, open_file, ys_conver_jz_excel
from API.GuiExec.gui_exec import user_login_role1, user_login_role0, update_combo_user1, update_combo_user2, \
    user_loginout
from API.MakeWindow import make_window as mw
from API.SqlExec.sql_exec import *
from CONFIG.settings import header, header1,file_dict,depart_dict,document_dict,approve_file_para
import PySimpleGUI as sg

# python -m pysimplegui-exemaker.pysimplegui-exemaker


if __name__ == '__main__':
    # 初始化变量
    login_user = ''
    addfile_filename = ''
    list_user = []  # 所有审批人
    wait_user = ''  # 待审批人
    list_id = []
    job_id = 0
    approve_job_id = 0  # 审批详情界面任务ID
    sql = ''
    result_sql = []
    max_number = 0
    query_job_count = 0
    new_number = '001'
    list_table = []  # 审批
    list_table3 = []
    list_table_lj = []  # 临技
    list_table_user = []  # 用户
    list_table_db = []
    list_ljtable = []
    list_table_wait_approve = []
    list_filename = []  # 添加的文件名
    db_table = ''
    department_db_cn = ''
    file_code_db = ''
    file_name_db = ''
    remark_info = ''
    pop_times = 1
    frame1_query_flag = 0
    frame3_query_flag = 0
    frame5_query_flag = 0
    ljtable_query_flag = 0
    double_event_flag = 0
    file_review_flag = 0
    temp_release_flag = 0
    my_send_flag = 0
    wait_approve_flag = 0
    finish_approve_flag = 0
    check_job_flag = 0
    check_popup_flag = 0

    # 表格初始化
    table = Texttable()
    table1 = Texttable()
    table.set_deco(Texttable.HEADER | Texttable.BORDER | Texttable.HLINES)  # |Texttable.HLINES|Texttable.VLINES
    table.set_cols_dtype(['i', 'a', 'a', 'a', 'a', 'a', 'a', 'a'])  # 列的数据类型
    table.set_header_align(["c", "c", "c", "c", "c", "c", "c", "c"])  # header的对齐方式
    table.set_cols_align(["c", "c", "c", "c", "c", "c", "c", "c"])  # 列的对齐方式
    table.set_cols_width([5, 20, 60, 8, 15, 15, 10, 12])  # 设置列的宽度
    table1.set_deco(Texttable.HEADER | Texttable.BORDER | Texttable.HLINES)
    table1.set_cols_dtype(['i', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a', 'a'])
    table1.set_header_align(["c", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c"])
    table1.set_cols_align(["c", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c", "c"])
    table1.set_cols_width([4, 16, 20, 4, 10, 10, 10, 4, 4, 4, 6, 10, 6, 10, 6])


    # 初始化窗口
    window2 = None
    window3 = None
    window4 = None
    window_login = None
    window_approve = None
    window_look_approve = None
    window_add_approve = None
    window_edit_approve = None
    window_make_sure = None
    window_add_user = None
    window_edit_user = None
    window_display_user_content = None
    window_online_user = None
    window_edit_user_info = None
    window_lj_info = None
    window_mes_ljtable = None
    window_mes_add = None
    window_mes_edit = None
    sec_window = None

    # 实例化主窗口
    sg.theme('LightGrey1')  # LightGrey1 LightGreen
    window1 = mw.make_window1()
    # 系统托盘
    menu_def = ['BLANK', ['&主界面', '---', '退出']]
    tray = sg.SystemTray(menu=menu_def, data_base64=sg.DEFAULT_BASE64_ICON, tooltip='极智文件系统')

    window1['FRAME1'].update(visible=False)
    window1['FRAME2'].update(visible=False)
    window1['FRAME3'].update(visible=False)
    window1['FRAME4'].update(visible=True)

    while True:
        window, event, values = sg.read_all_windows(timeout=1000)

        # frame1 绑定回车事件
        window1['DEPART'].bind("<Return>", 'Return')
        window1['FILE_CODE_INPUT'].bind("<Return>", 'Return')
        window1['FILE_NAME_INPUT'].bind("<Return>", 'Return')
        window1['REMARK_INPUT'].bind("<Return>", 'Return')

        # frame3 绑定回车事件
        window1['DEPART_DB'].bind("<Return>", 'Return')
        window1['FILE_CODE_DB_INPUT'].bind("<Return>", 'Return')
        window1['FILE_NAME_DB_INPUT'].bind("<Return>", 'Return')
        window1['REMARK_DB_INPUT'].bind("<Return>", 'Return')

        # frame5 绑定回车事件
        window1['FILE_CODE_LJ'].bind("<Return>", 'Return')
        window1['PRO_CODE_LJ'].bind("<Return>", 'Return')
        window1['RWL_LJ'].bind("<Return>", 'Return')
        window1['REMARK_LJ'].bind("<Return>", 'Return')
        window1['TYPE_LJ'].bind("<Return>", 'Return')
        window1['TIME_LJ'].bind("<Return>", 'Return')

        # 绑定双击事件
        window1['-TABLE-'].bind("<Double-Button-1>", 'Double')
        window1['TABLE2'].bind("<Double-Button-1>", 'Double')
        window1['TABLE-LJ'].bind("<Double-Button-1>", 'Double')


        if window == window1 and event is None:
            window1.hide()
            tray.un_hide()

        if event == '主界面' or event == '-IMAGE-':
            window1.un_hide()

        if event == '退出':
            tray.close()
            break

        if event == '文件查询':
            window1['FRAME1'].update(visible=True)
            window1['FRAME2'].update(visible=False)
            window1['FRAME3'].update(visible=False)
            window1['FRAME4'].update(visible=False)
            window1['FRAME5'].update(visible=False)

        if event == '二维码':
            window1['FRAME1'].update(visible=False)
            window1['FRAME2'].update(visible=True)
            window1['FRAME3'].update(visible=False)
            window1['FRAME4'].update(visible=False)
            window1['FRAME5'].update(visible=False)

        if event == '数据库':
            window1['FRAME1'].update(visible=False)
            window1['FRAME2'].update(visible=False)
            window1['FRAME3'].update(visible=True)
            window1['FRAME4'].update(visible=False)
            window1['FRAME5'].update(visible=False)

        if event == '审批申请':
            window1['FRAME1'].update(visible=False)
            window1['FRAME2'].update(visible=False)
            window1['FRAME3'].update(visible=False)
            window1['FRAME4'].update(visible=True)
            window1['FRAME5'].update(visible=False)

        if event == '临技查询':
            window1['FRAME1'].update(visible=False)
            window1['FRAME2'].update(visible=False)
            window1['FRAME3'].update(visible=False)
            window1['FRAME4'].update(visible=False)
            window1['FRAME5'].update(visible=True)

        if event in ('取消', sg.WIN_CLOSED, 'Exit'):
            if window == window1:
                window1.hide()
                tray.un_hide()
            elif window == window_display_user_content:
                window_display_user_content.close()
            elif window == window_edit_user_info:
                window_edit_user_info.close()
            elif window == window_mes_ljtable:
                window_mes_ljtable.close()
            else:
                sec_window.close()
                sec_window = None

        if event in ('DEPARTReturn', 'FILE_CODE_INPUTReturn', 'FILE_NAME_INPUTReturn', 'REMARK_INPUTReturn'):
            frame1_query_flag = 1

        if event in (
                'DEPART_DBReturn', 'FILE_CODE_DB_INPUTReturn', 'FILE_NAME_DB_INPUTReturn', 'REMARK_DB_INPUTReturn'):
            frame3_query_flag = 1

        if event in ('FILE_CODE_LJReturn', 'PRO_CODE_LJReturn', 'RWL_LJReturn', 'REMARK_LJReturn',
                     'TYPE_LJReturn', 'TIME_LJReturn'):
            frame5_query_flag = 1

        if event in ('MES-RWL-INPUTReturn', 'MES-CODE-INPUTReturn', 'MES-REMARK-INPUTReturn'):
            ljtable_query_flag = 1

        if event == '-TABLE-Double':
            if login_user == '':
                sg.popup_auto_close('登录后才能操作', title='请登录', font=('宋体', 18))
            else:
                double_event_flag = 1

        if event == 'LOGIN-BUTTON':
            if sec_window is not None:
                sec_window.close()
            window_login = mw.make_window_login()
            sec_window = window_login
            window_login['LOGIN-USER'].bind("<Return>", 'Return')
            window_login['LOGIN-PASSWD'].bind("<Return>", 'Return')

        if event == 'LOGIN' or event == 'LOGIN-PASSWDReturn':
            login_user = values['LOGIN-USER']
            login_passwd = values['LOGIN-PASSWD']
            sql = f"SELECT id_password,role from member WHERE id_number='{login_user}' "
            result_sql = exec_sql_server_utf8(sql)
            print(result_sql)
            if len(result_sql) != 0:
                passwd_md5 = computerMD5(login_passwd)
                if passwd_md5 == result_sql[0][0]:
                    write_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                    user_name = exec_sql_server_utf8(f"SELECT name FROM member WHERE id_number='{login_user}' ")
                    window1['LOGIN-TEXT'].update(f'用户:{login_user}-{user_name[0][0]}')
                    window1['USER-PNG'].update(filename=user_png)
                    sec_window.close()
                    insert_para = '(id_number,login_time,login_status)'
                    insert_value = (login_user, write_time, '1')
                    sql = f"INSERT INTO login_info {insert_para} VALUES {insert_value} "
                    update_sql_server(sql)
                    update_sql_server(f"UPDATE member SET user_status=1 WHERE id_number='{login_user}' ")
                    if result_sql[0][1] == 1:
                        user_login_role1(window1)
                    elif result_sql[0][1] == 2:
                        user_login_role1(window1)
                    else:
                        user_login_role0(window1)
                    logger.info(f"{login_user} 登录成功")
                    wait_approve_flag = 1  # 查询待审批文件
                    file_review_flag = 1  # 显示审批流程
                    check_job_flag = 1  # 检查任务弹屏
                    check_popup_flag = 1  # 开启审批完成弹屏

                else:
                    sg.popup_auto_close('登录失败,用户或密码错误', title='登录失败', font=('宋体', 18))
            else:
                sg.popup_auto_close('登录失败,用户无效', title='登录失败', font=('宋体', 18))

        if event == 'LOGIN-OUT':
            write_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
            user_loginout(window1)
            check_job_flag = 0
            check_popup_flag = 0
            insert_para = '(id_number,login_time,login_status)'
            insert_value = (login_user, write_time, '0')
            sql = f"INSERT INTO login_info {insert_para} VALUES {insert_value} "
            update_sql_server(sql)
            update_sql_server(f"UPDATE member SET user_status=0 WHERE id_number='{login_user}' ")
            logger.info(f"{login_user} 退出登录")

        if event == 'FRAME1_QUERY' or frame1_query_flag == 1:
            frame1_query_flag = 0
            window['-FILE_SYSTEM-'].update('')
            result_sql = []
            filetype = values['FILE_TYPE']
            department = values['DEPART']
            filecode = values['FILE_CODE_INPUT']
            filename = values['FILE_NAME_INPUT']
            remark_info = values['REMARK_INPUT']
            data_table = file_dict.get(filetype, '')
            print(f'文件类型:{filetype}')
            print(f'部门:{department}')
            print(f'文件编号:{filecode}')
            print(f'文件名称:{filename}')
            print(f'表名:{data_table}')
            if data_table != '':
                if filetype == '文件发放':
                    sql = document_issue_sql(department, filecode, filename, remark_info)

                elif filetype == '所有文件':
                    sql = all_table_sql(department, filecode, filename, remark_info)

                else:
                    sql = other_table_sql(data_table, department, filecode, filename, remark_info)

                result_sql = exec_sql_server_utf8(sql)
                if data_table == 'document_issue':
                    result_sql.insert(0, header1)  # header插入到开头
                    table1.add_rows(result_sql)  # 表增加列,
                    sg.cprint(table1.draw())
                    table1.reset()
                else:
                    result_sql.insert(0, header)  # header插入到开头
                    table.add_rows(result_sql)  # 表增加列,
                    sg.cprint(table.draw())
                    table.reset()
            else:
                sg.popup_error('第一个选择框不能为空或有误,请重新选择', title='错误', font=('宋体', 18))

        if event == 'GENERATE_CODE':  # 编号申请
            window['-FILE_SYSTEM-'].update('')
            filetype = values['FILE_TYPE']
            department = values['DEPART']
            data_table = file_dict[f'{filetype}']
            department_en = depart_dict[f'{department}']
            wl_file_code = values['WL_FILE_CODE']
            date = time.strftime('%Y-%m-%d', time.localtime(time.time()))

            if filetype == '二阶文件':
                sql = f"SELECT ID,file_code,file_name,version,createdate,department,remark FROM level2_document " \
                      f"order by ID "
                sql_maxvalue = f"SELECT MAX(ID) FROM level2_document"
                res = exec_sql_server_utf8(sql)
                new_id = table_max_id('level2_document')
                if len(res) == 0:
                    new_number = '001'
                else:
                    item = res[-1][1]
                    print(item)
                    item_split = item.split('-')
                    if len(item_split) == 3:
                        number = int(item_split[2])
                        new_number = number + 1
                        if new_number < 100:
                            new_number = str(new_number).zfill(3)
                        else:
                            new_number = str(new_number)
                    new_filecode = 'JZ-' + document_dict[f'{filetype}'] + '-' + new_number
                    print(f'申请的ID,新编号:{new_id},{new_filecode}')
                    logger.info(f'申请的ID,新编号:{new_id},{new_filecode},申请人:{login_user}')
                    # window["-FILE_SYSTEM-"].update(f'新申请的文件编号:{new_filecode}')
                    sg.cprint(f'新申请的文件编号:{new_filecode}', colors='red')
                    insert_para = '(ID,file_code,createdate,department,file_type,remark)'
                    insert_value = (
                        new_id, new_filecode, date, department, filetype, login_user)
                    insert_sql = f"INSERT {data_table} {insert_para} VALUES {insert_value}"
                    update_sql_server(insert_sql)

            if filetype == '三阶文件' or filetype == '四阶文件':
                if department == '':
                    sg.popup_auto_close('部门不能为空，请选择部门', title='注意', font=('宋体', 18))
                else:
                    sql = f"SELECT ID,file_code,file_name,version,createdate,department,remark FROM {data_table} " \
                          f"where department LIKE '%{department}%' order by ID"
                    sql_maxvalue = f"SELECT MAX(ID) FROM {data_table}"
                    res = exec_sql_server_utf8(sql)
                    res_maxvalue = exec_sql_server(sql_maxvalue)
                    new_id = table_max_id(data_table)

                    if len(res) == 0:
                        new_number = '001'
                    else:
                        item = res[-1][1]
                        print(item)
                        item_split = item.split('-')
                        if len(item_split) == 4:
                            number = int(item_split[3])
                            new_number = number + 1
                            if new_number < 100:
                                new_number = str(new_number).zfill(3)
                            else:
                                new_number = str(new_number)
                    new_filecode = 'JZ-' + document_dict[f'{filetype}'] + '-' + department_en + '-' + new_number
                    # window["-FILE_SYSTEM-"].update(f'新申请的文件编号:{new_filecode}')
                    sg.cprint(f'新申请的文件编号:{new_filecode}', colors='red')
                    print(f'申请的ID,新编号:{new_id},{new_filecode}, 申请人:{login_user}')
                    logger.info(f'申请的ID,新编号:{new_id},{new_filecode}')
                    if filetype == '三阶文件':
                        insert_para = '(ID,file_code,createdate,department,file_type,remark)'
                        insert_value = (
                            new_id, new_filecode, date, department, filetype, login_user)
                        insert_sql = f"INSERT {data_table} {insert_para} VALUES {insert_value}"
                        update_sql_server(insert_sql)
                    else:
                        insert_para = '(ID,file_code,createdate,department,file_type,remark)'
                        insert_value = (
                            new_id, new_filecode, date, department, filetype, login_user)
                        insert_sql = f"INSERT {data_table} {insert_para} VALUES {insert_value}"
                        update_sql_server(insert_sql)

            if filetype == '外来文件':
                if wl_file_code == '':
                    sg.popup_auto_close('外来文件客户代码不能为空', title='注意', font=('宋体', 18))
                else:
                    temp_number = 1
                    year = time.strftime('%Y', time.localtime(time.time()))
                    month = time.strftime('%m', time.localtime(time.time()))
                    sql = f"SELECT file_code FROM {data_table}"
                    sql_maxvalue = f"SELECT MAX(ID) FROM {data_table}"
                    result_sql = exec_sql_server(sql)
                    res_maxvalue = exec_sql_server(sql_maxvalue)
                    new_id = table_max_id(data_table)
                    for i in range(len(result_sql)):
                        item = result_sql[i][0]
                        item_list = item.split('-')
                        if item_list[1] == year and item_list[2] == month:
                            if int(item_list[3]) >= temp_number:
                                max_number = int(item_list[3])
                        else:
                            max_number = 0
                    new_number = max_number + 1
                    if new_number < 100:
                        new_number = str(new_number).zfill(3)
                    else:
                        new_number = str(new_number)
                    new_filecode = wl_file_code + 'WL-' + year + '-' + month + '-' + new_number
                    # window["-FILE_SYSTEM-"].update(f'新申请的文件编号:{new_filecode}')
                    sg.cprint(f'新申请的文件编号:{new_filecode}', colors='red')
                    print(f'申请的ID,新编号:{new_id},{new_filecode}')
                    logger.info(f'申请的ID,新编号:{new_id},{new_filecode}, 申请人:{login_user}')
                    insert_para = '(ID,file_code,createdate,department,file_type,remark)'
                    insert_value = (new_id, new_filecode, date, department, filetype, login_user)
                    insert_sql = f"INSERT {data_table} {insert_para} VALUES {insert_value}"
                    update_sql_server(insert_sql)

            if filetype == '临时文件':
                temp_number = 1
                year = time.strftime('%Y', time.localtime(time.time()))
                month = time.strftime('%m', time.localtime(time.time()))
                sql = f"SELECT file_code FROM {data_table}"
                sql_maxvalue = f"SELECT MAX(ID) FROM {data_table}"
                result_sql = exec_sql_server(sql)
                res_maxvalue = exec_sql_server(sql_maxvalue)
                new_id = table_max_id(data_table)
                for i in range(len(result_sql)):
                    item = result_sql[i][0]
                    item_list = item.split('-')
                    if item_list[1] == year and item_list[2] == month:
                        if int(item_list[3]) >= temp_number:
                            max_number = int(item_list[3])
                            print('max_nubmer:', max_number)
                    else:
                        max_number = 0
                new_number = max_number + 1
                if new_number < 100:
                    new_number = str(new_number).zfill(3)
                else:
                    new_number = str(new_number)
                new_filecode = 'LS-' + year + '-' + month + '-' + new_number
                sg.cprint(f'新申请的文件编号:{new_filecode}', colors='red')
                print(f'申请的ID,新编号:{new_id},{new_filecode}')
                logger.info(f'申请的ID,新编号:{new_id},{new_filecode}, 申请人:{login_user}')
                insert_para = '(ID,file_code,createdate,department,file_type,remark)'
                insert_value = (new_id, new_filecode, date, department, filetype, login_user)
                insert_sql = f"INSERT {data_table} {insert_para} VALUES {insert_value}"
                update_sql_server(insert_sql)

            if filetype == '文件发放':
                sg.popup_auto_close('文件发放不能申请编号,请选择其他文件', title='注意', font=('宋体', 18))

            if filetype == '所有文件':
                sg.popup_auto_close('所有文件不能申请编号,请选择其他文件', title='注意', font=('宋体', 18))

        if event == 'FILE_OUTPUT':
            filetype = values['FILE_TYPE']
            now_time = time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime(time.time()))
            new_excel = excel_path + '\\' + filetype + f'-{now_time}.xlsx'
            print(new_excel)
            wb = Workbook()  # 创建工作薄对象
            ws = wb.create_sheet(f'{filetype}', 0)  # 创建工作表对象,0对应第1个sheet，1为第2个sheet
            if len(result_sql) != 0:
                try:
                    for i in range(len(result_sql)):
                        item = result_sql[i]
                        print(item)
                        write_row_by_list(ws, i + 1, item)
                    wb.save(new_excel)
                    logger.info(f'导出成功,路径:{new_excel}')
                    sg.popup_ok(f'路径:{new_excel}', title='导出成功', font=('宋体', 18), )

                except Exception as result:
                    sg.popup_error(result, font=('宋体', 18))
            else:
                sg.popup_auto_close('请先查询，再导出', title='注意', font=('宋体', 18))

        if event == 'OPEN-DIR':
            try:
                os.startfile(excel_path)
                logger.info(f'打开文件夹{excel_path}')
            except Exception as e:
                sg.popup_error(e, title='错误', font=('宋体', 18))

        if event == 'OPEN-QR-DIR':
            try:
                os.startfile(png_path)
                logger.info(f'打开文件夹{png_path}')
            except Exception as e:
                sg.popup_error(e, font=('宋体', 18))

        if event == '-MAKE_QRCODE-':  # 生成二维码
            input_qrcode = values['-INPUT_QRCODE-']
            file_name = input_qrcode + '.png'
            img_path = fr'{png_path}\{file_name}'
            print(img_path)
            make_qr_code(input_qrcode, img_path)
            logger.info(f'二维码内容:{input_qrcode}')
            logger.info(f'二维码路径:{img_path}')
            window["QRCODE_PNG_PATH"].update(f'{img_path}')
            window["-IMAGE-"].update(filename=img_path, size=(1440, 960))

        if event == 'FRAME3_QUERY' or frame3_query_flag == 1:
            frame3_query_flag = 0
            list_table_db = []
            file_type_db = values['FILE_TYPE_DB']
            department = values['DEPART_DB']
            filename = values['FILE_NAME_DB_INPUT']
            filecode = values['FILE_CODE_DB_INPUT']
            remark_info = values['REMARK_DB_INPUT']
            db_table = file_dict.get(file_type_db, '')
            if db_table != '':
                if file_type_db == '所有文件':
                    flush_all_display(department, filecode, filename, remark_info)

                else:
                    flush_display(db_table, department, filecode, filename, remark_info)
            else:
                sg.popup_error('第一个选择框不能为空或有误,请重新选择', title='错误', font=('宋体', 18))

        if event == 'DEL-BUTTON':
            res = 0
            count = 0
            list_id_value = []
            file_type_db = values['FILE_TYPE_DB']
            department_db_cn = values['DEPART_DB']
            file_code_db = values['FILE_CODE_DB_INPUT']
            file_name_db = values['FILE_NAME_DB_INPUT']
            remark_info = values['REMARK_DB_INPUT']
            db_table = file_dict[f'{file_type_db}']
            select_row = values['-TABLE-']
            if len(select_row) == 0:
                sg.popup_auto_close('请选择需要删除的行', title='注意', font=('宋体', 18))
            else:
                for i in range(len(select_row)):
                    print(f'选择的行:{select_row[i]}')
                    select_value = list_table_db[select_row[i]]
                    print(select_value)
                    id_item = select_value[0]
                    table_name_cn = select_value[6]
                    db_table_del = file_dict[f'{table_name_cn}']

                    res = update_sql_server(f"DELETE {db_table_del} WHERE ID={id_item}")
                    if res == 1:
                        list_id_value.append(id_item)
                        logger.info(f"删除{db_table_del} ID为{id_item}的行成功")
                        logger.info(f'删除内容:{select_value}')
                        count += 1
                    else:
                        logger.info(f"删除{db_table_del} ID为{id_item}的行失败")
                        sg.popup_auto_close('删除失败', title='失败', font=('宋体', 18))
                if db_table == 'all_table':
                    flush_all_display(department_db_cn, file_code_db, file_name_db, remark_info)
                else:
                    flush_display(db_table, department_db_cn, file_code_db, file_name_db, remark_info)
                if len(select_row) == count:
                    sg.popup_auto_close(f'ID {list_id_value} 删除成功', title='成功', font=('宋体', 18))

        if event == 'MOD-BUTTON' or double_event_flag == 1:
            double_event_flag = 0
            file_type_db = values['FILE_TYPE_DB']
            department_db_cn = values['DEPART_DB']
            file_code_db = values['FILE_CODE_DB_INPUT']
            file_name_db = values['FILE_NAME_DB_INPUT']
            remark_info = values['REMARK_DB_INPUT']
            db_table = file_dict.get(file_type_db, '')
            select_row = values['-TABLE-']
            if len(select_row) == 0:
                sg.popup_auto_close('请选择需要修改的行', title='注意', font=('宋体', 18))
            else:
                if sec_window is not None:
                    sec_window.close()
                window3 = mw.make_window3()
                sec_window = window3
                window3['MOD-INPUT'].bind("<Return>", 'Return')

                if len(list_table_db) > 0:
                    select_item = list_table_db[select_row[0]]
                    print(select_item)
                    if len(select_item) > 0:
                        item_id = select_item[0]
                        item_filecode = select_item[1]
                        item_filename = select_item[2]
                        item_version = select_item[3]
                        item_time = select_item[4]
                        item_depart = select_item[5]
                        item_type = select_item[6]
                        item_remark = select_item[7]
                        window3['MOD-ID'].update(item_id)
                        window3['MOD-CODE'].update(item_filecode)
                        window3['MOD-NAME'].update(item_filename)
                        window3['MOD-VER'].update(item_version)
                        window3['MOD-TIME'].update(item_time)
                        window3['MOD-DEP'].update(item_depart)
                        window3['MOD-TYPE'].update(item_type)
                        window3['MOD-REMARK'].update(item_remark)

        if event == 'MOD-UPDATE' or event == 'MOD-INPUTReturn':
            mod_input = values['MOD-INPUT']
            file_code_input = values['MOD-CODE']
            input_split = mod_input.split('_')
            print(input_split)
            if len(input_split) == 4:
                file_code = input_split[0]
                version = input_split[1]
                file_name = input_split[2]
                page_number = input_split[3]

                if file_code == file_code_input:
                    window3['MOD-NAME'].update(file_name)
                    window3['MOD-VER'].update(version)
                else:
                    sg.popup_auto_close('文件编号不一致，请检查', title='注意', font=('宋体', 18))
            else:
                sg.popup_auto_close('二维码输入格式有误，请检查', title='错误', font=('宋体', 18))

        if window == window3 and event == 'MOD-SAVE':
            # print(values)
            file_type_cn = window1['FILE_TYPE_DB'].get()
            item_id = values['MOD-ID']
            item_filecode = values['MOD-CODE']
            item_filename = values['MOD-NAME']
            item_version = values['MOD-VER']
            item_time = values['MOD-TIME']
            item_depart = values['MOD-DEP']
            item_type = values['MOD-TYPE']
            item_remark = values['MOD-REMARK']
            db_table_new = file_dict.get(item_type, '')
            sql_diff_filename = f"SELECT * FROM {db_table_new} WHERE file_name='{item_filename}'"
            res_diff = exec_sql_server_utf8(sql_diff_filename)
            print(res_diff)
            if len(res_diff) == 0:
                sql = f"UPDATE {db_table_new} SET file_code='{item_filecode}',file_name='{item_filename}',version='{item_version}' ," \
                      f"createdate='{item_time}', department='{item_depart}',remark='{item_remark}' " \
                      f"WHERE ID={item_id}"

            if len(res_diff) > 0:
                res_id = str(res_diff[0][0])
                print(f'item_id:{item_id}, res_id:{res_id}')
                if item_id == res_id:
                    sql = f"UPDATE {db_table_new} SET file_code='{item_filecode}',file_name='{item_filename}',version='{item_version}' ," \
                          f"createdate='{item_time}', department='{item_depart}',remark='{item_remark}' " \
                          f"WHERE ID={item_id}"

                else:
                    sg.popup_auto_close('文件名称有重名，请检查', title='错误', font=('宋体', 18))

            res = update_sql_server(sql)
            if res == 1:
                window3.close()
                if file_type_cn == '所有文件':
                    flush_all_display(department_db_cn, file_code_db, file_name_db, remark_info)
                else:
                    flush_display(db_table_new, department_db_cn, file_code_db, file_name_db, remark_info)
                logger.info('更新数据成功')
                sg.popup_auto_close('更新数据成功', title='成功', font=('宋体', 18))
            else:
                window3.close()
                logger.info('更新数据失败')
                sg.popup_auto_close('更新数据失败', title='失败', font=('宋体', 18))

        if event == 'QR-UPDATE-BUTTON':
            file_type_db = values['FILE_TYPE_DB']
            department_db_cn = values['DEPART_DB']
            file_code_db = values['FILE_CODE_DB_INPUT']
            file_name_db = values['FILE_NAME_DB_INPUT']
            remark_info = values['REMARK_DB_INPUT']
            if sec_window is not None:
                sec_window.close()
            window2 = mw.make_window2()
            sec_window = window2
            window2['QR-INPUT'].bind("<Return>", 'Return')
            res_max_id = exec_sql_server("SELECT MAX(ID) FROM document_issue")
            new_id = int(res_max_id[0][0]) + 1
            window2['QR-ID'].update(new_id)

        if event == 'QR-UPDATE' or event == 'QR-INPUTReturn':
            # print('第2个窗口对象:', sec_window)
            # print(values)
            qrcode_input = values['QR-INPUT']
            department = values['QR-DEP']
            receipt_user = values['QR-USER']
            remark = values['QR-REMARK']
            number = 1
            input_split = qrcode_input.split('_')
            print(input_split)
            if len(input_split) == 4:
                file_code = input_split[0]
                version = input_split[1]
                file_name = input_split[2]
                page_number = input_split[3]

                date = time.strftime('%Y-%m-%d', time.localtime(time.time()))
                window2['QR-NAME'].update(file_name)
                window2['QR-CODE'].update(file_code)
                window2['QR-VER'].update(version)
                window2['QR-USER'].update(receipt_user)
                window2['QR-NUM'].update(number)
                window2['QR-DATE'].update(date)
                window2['QR-PAGE'].update(page_number)
                window2['QR-REMARK'].update(remark)
            else:
                sg.popup_auto_close('二维码输入格式有误，请检查', title='错误', font=('宋体', 18))

        if window == window2 and event == 'QR-SAVE':
            write_id = values['QR-ID']
            file_name = values['QR-NAME']
            file_code = values['QR-CODE']
            version = values['QR-VER']
            department = values['QR-DEP']
            receipt_user = values['QR-USER']
            number = values['QR-NUM']
            date = values['QR-DATE']
            page_number = values['QR-PAGE']
            file_flag = values['QR-FLAG']
            file_type = '文件发放'
            remark = values['QR-REMARK']
            recover_status = ''
            recover_date = ''
            invalid_status = ''
            invalid_date = ''
            insert_value = (write_id, file_name, file_code, version, department, receipt_user,
                            date, number, page_number, file_flag, recover_status, recover_date,
                            invalid_status, invalid_date, file_type, remark)
            res = update_sql_server(f"INSERT INTO document_issue VALUES {insert_value}")
            if res == 1:
                window2.close()
                window1['FILE_TYPE_DB'].update('文件发放')
                flush_display('document_issue', department_db_cn, file_code_db, file_name_db, remark_info)
                logger.info('写入文件发放成功')
                sg.popup_auto_close('写入成功', title='成功', font=('宋体', 18))
            else:
                window2.close()
                logger.info('写入文件发放失败')
                sg.popup_auto_close('写入失败', title='失败', font=('宋体', 18))

        if event == 'CIRCLE-BUTTON':
            file_type_db = values['FILE_TYPE_DB']
            department_db_cn = values['DEPART_DB']
            file_code_db = values['FILE_CODE_DB_INPUT']
            file_name_db = values['FILE_NAME_DB_INPUT']
            remark_info = values['REMARK_DB_INPUT']
            select_row = values['-TABLE-']
            if file_type_db == '文件发放':
                if len(select_row) == 0:
                    sg.popup_auto_close('请选择需要修改的行', title='注意', font=('宋体', 18))
                else:
                    row_number = select_row[0]
                    row_value = list_table_db[row_number]
                    id_value = row_value[0]
                    date = time.strftime('%Y-%m-%d', time.localtime(time.time()))
                    sql = f"UPDATE document_issue SET recover_status='已回收', recoverdate='{date}', " \
                          f"remark='已回收' WHERE ID={id_value}"
                    res = update_sql_server(sql)
                    if res == 1:
                        flush_display('document_issue', department_db_cn, file_code_db, file_name_db, remark_info)
                        logger.info(f'ID:{id_value} 文件回收成功')
                        sg.popup_auto_close('操作成功', title='成功', font=('宋体', 18))
                    else:
                        logger.info(f'ID:{id_value} 文件回收失败')
                        sg.popup_auto_close('操作失败', title='失败', font=('宋体', 18))

            else:
                sg.popup_auto_close('文件回收请选择文件发放', title='注意', font=('宋体', 18))

        if event == 'DISABLE-BUTTON':
            file_type_db = values['FILE_TYPE_DB']
            department_db_cn = values['DEPART_DB']
            file_code_db = values['FILE_CODE_DB_INPUT']
            file_name_db = values['FILE_NAME_DB_INPUT']
            remark_info = values['REMARK_DB_INPUT']
            select_row = values['-TABLE-']
            if file_type_db == '文件发放':
                if len(select_row) == 0:
                    sg.popup_auto_close('请选择需要修改的行', title='注意', font=('宋体', 18))
                else:
                    row_number = select_row[0]
                    row_value = list_table_db[row_number]
                    id_value = row_value[0]
                    date = time.strftime('%Y-%m-%d', time.localtime(time.time()))
                    sql = f"UPDATE document_issue SET invalid_status='已作废', invaliddate='{date}', " \
                          f"remark='已作废' WHERE ID={id_value}"
                    res = update_sql_server(sql)
                    if res == 1:
                        flush_display('document_issue', department_db_cn, file_code_db, file_name_db, remark_info)
                        logger.info(f'ID:{id_value} 文件作废成功')
                        sg.popup_auto_close('操作成功', title='成功', font=('宋体', 18))
                    else:
                        logger.info(f'ID:{id_value} 文件作废失败')
                        sg.popup_auto_close('操作失败', title='失败', font=('宋体', 18))

            else:
                sg.popup_auto_close('文件作废请选择文件发放', title='注意', font=('宋体', 18))

        if event == 'INFO-BUTTON':
            file_type_db = values['FILE_TYPE_DB']
            department_db_cn = values['DEPART_DB']
            file_code_db = values['FILE_CODE_DB_INPUT']
            file_name_db = values['FILE_NAME_DB_INPUT']
            remark_info = values['REMARK_DB_INPUT']
            select_row = values['-TABLE-']
            if file_type_db == '文件发放':
                if len(select_row) == 0:
                    sg.popup_auto_close('请选择需要修改的行', title='注意', font=('宋体', 18))
                else:
                    if sec_window is not None:
                        sec_window.close()
                    window4 = mw.make_window4()
                    sec_window = window4
                    row_number = select_row[0]
                    row_value = list_table_db[row_number]
                    id_value = row_value[0]
                    sql = f"SELECT * FROM document_issue WHERE ID={id_value}"
                    res = exec_sql_server_utf8(sql)
                    if len(res) != 0:
                        info_id = res[0][0]
                        info_name = res[0][1]
                        info_code = res[0][2]
                        info_ver = res[0][3]
                        info_dep = res[0][4]
                        info_user = res[0][5]
                        info_rdate = res[0][6]
                        info_num = res[0][7]
                        info_page = res[0][8]
                        info_type = res[0][9]
                        info_recover = res[0][10]
                        info_hdate = res[0][11]
                        info_invalid = res[0][12]
                        info_idate = res[0][13]
                        info_remark = res[0][14]
                        # 页面显示
                        window4['INFO-ID'].update(info_id)
                        window4['INFO-NAME'].update(info_name)
                        window4['INFO-CODE'].update(info_code)
                        window4['INFO-VER'].update(info_ver)
                        window4['INFO-DEP'].update(info_dep)
                        window4['INFO-USER'].update(info_user)
                        window4['INFO-RDATE'].update(info_rdate)
                        window4['INFO-NUM'].update(info_num)
                        window4['INFO-PAGE'].update(info_page)
                        window4['INFO-TYPE'].update(info_type)
                        window4['INFO-RECOVER'].update(info_recover)
                        window4['INFO-HDATE'].update(info_hdate)
                        window4['INFO-INVALID'].update(info_invalid)
                        window4['INFO-IDATE'].update(info_idate)
                        window4['INFO-REMARK'].update(info_remark)

            else:
                sg.popup_auto_close('详情请选择文件发放', title='注意', font=('宋体', 18))

        if window == window4 and event == 'INFO-UPDATE':
            info_id = values['INFO-ID']
            info_name = values['INFO-NAME']
            info_code = values['INFO-CODE']
            info_ver = values['INFO-VER']
            info_dep = values['INFO-DEP']
            info_user = values['INFO-USER']
            info_rdate = values['INFO-RDATE']
            info_num = values['INFO-NUM']
            info_page = values['INFO-PAGE']
            info_type = values['INFO-TYPE']
            info_recover = values['INFO-RECOVER']
            info_hdate = values['INFO-HDATE']
            info_invalid = values['INFO-INVALID']
            info_idate = values['INFO-IDATE']
            info_remark = values['INFO-REMARK']
            print(info_id, info_code, info_name, info_ver, info_dep, info_user, info_rdate, info_num,
                  info_page, info_type, info_recover, info_hdate, info_invalid, info_idate, info_remark)
            sql = f"UPDATE document_issue set file_name='{info_name}',file_code='{info_code}',version='{info_ver}'," \
                  f"department='{info_dep}', receipt_user='{info_user}', createdate='{info_rdate}'," \
                  f"number='{info_num}', page_number='{info_page}', file_flag='{info_type}'," \
                  f"recover_status='{info_recover}', recoverdate='{info_hdate}', invalid_status='{info_invalid}'," \
                  f"invaliddate='{info_idate}',remark='{info_remark}' " \
                  f"where ID={info_id}"
            res = update_sql_server(sql)
            if res == 1:
                flush_display('document_issue', department_db_cn, file_code_db, file_name_db, remark_info)
                window4.close()
                logger.info('更新文件发放成功')
                sg.popup_auto_close('保存成功', title='成功', font=('宋体', 18))
            else:
                window4.close()
                logger.info('更新文件发放失败')
                sg.popup_auto_close('保存失败', title='失败', font=('宋体', 18))

        if event == '打开文件夹':
            direction = ''
            select_row = values['-TABLE-']
            if len(select_row) == 0:
                sg.popup_auto_close('选择不能为空', title='注意', font=('宋体', 18))
            elif len(select_row) > 1:
                sg.popup_auto_close('打开文件时，不能选择多行', title='注意', font=('宋体', 18))
            else:
                select_value = list_table_db[select_row[0]]
                doc_code = select_value[1]
                doc_dep = select_value[5]
                doc_type = select_value[6]
                print(f'文件编号-部门-文件类型:{doc_code}-{doc_dep}-{doc_type}')
                if doc_type == '二阶文件':
                    direction = document_path + '\\' + '二阶文件(新)'
                elif doc_type == '临技文件':
                    direction = r'\\10.10.80.80\share\宇视\临技\已处理'
                else:
                    direction = document_path + '\\' + doc_type
                print('文件路径:', direction)
                try:
                    os.startfile(direction)
                    logger.info(f'打开文件夹 {direction}')
                except Exception as e:
                    sg.popup_error(e, font=('宋体', 18))

        if event == '打开文件':
            new_path = ''
            select_value = ''
            list_pe_dep = ['IE部', 'TE', '生产部SMT', '工程部', 'ME']
            dict_pe_dep = {'IE部': 'IE', 'TE': 'TE', '生产部SMT': 'SMT', '工程部': 'PE', 'ME': 'ME'}
            select_row = values['-TABLE-']
            if len(select_row) == 0:
                sg.popup_auto_close('请选择需要打开的文件', title='注意', font=('宋体', 18))
            elif len(select_row) > 1:
                sg.popup_auto_close('打开文件时，不能选择多行', title='注意', font=('宋体', 18))
            else:
                # print('list_table_db:', list_table_db)
                print('选择的行:', select_row)
                if len(list_table_db) > 0:
                    select_value = list_table_db[select_row[0]]
                doc_code = select_value[1]
                doc_name = select_value[2]
                doc_version = select_value[3]
                doc_dep = select_value[5]
                doc_type = select_value[6]
                if doc_name not in (None, ''):
                    if doc_version in (None, ''):
                        full_name = doc_code + '_' + doc_name
                    else:
                        doc_version = doc_version.replace(' ', '')
                        full_name = doc_code + '_' + doc_version + '_' + doc_name
                    print(f'文件编号-文件名称-部门-文件类型:{doc_code}-{doc_name}-{doc_dep}-{doc_type}')
                    res_server_path = exec_sql_server_utf8(f"SELECT server_path FROM approve_file "
                                                           f"WHERE file_name LIKE N'%{full_name}%' ")
                    if len(res_server_path) > 0:
                        server_path = res_server_path[0][0]
                        print(f'文件路径:{server_path}')
                        try:
                            if server_path not in ('', None):
                                # os.system(server_path)
                                cmd_param = 'start ' + server_path
                                myPopen(cmd_param)
                        except Exception as e:
                            sg.popup_error(e, title='错误', font=('宋体', 18))

                    else:
                        if doc_type == '二阶文件':
                            new_path = document_path + '\\' + '二阶文件(新)'
                        elif doc_type == '临技文件':
                            new_path = r'\\10.10.80.80\share\宇视\临技\已处理'
                        elif doc_type == '三阶文件' and doc_dep in list_pe_dep:
                            new_path = document_path + '\\' + '三阶文件' + '\\' + '工程部' + '\\' + dict_pe_dep[f'{doc_dep}']
                        else:
                            new_path = document_path + '\\' + doc_type

                        open_file(new_path, doc_code)
                else:
                    sg.popup_error('文件名称为空,无法打开', title='错误', font=('宋体', 18))

        if event == 'FILE-REVIEW' or file_review_flag == 1:
            file_review_flag = 0
            listbox_value1 = []
            sql = f"SELECT approve_name from approve_user WHERE file_type=1 ORDER BY approve_name"
            res = exec_sql_server_utf8(sql)
            for i in range(len(res)):
                item = res[i][0]
                listbox_value1.append(item)
            window1['LIST-BOX'].update(listbox_value1)

        if event == 'TEMP-RELEASE' or temp_release_flag == 1:
            temp_release_flag = 0
            listbox_value1 = []
            sql = f"SELECT approve_name from approve_user WHERE file_type=2 ORDER BY approve_name "
            res = exec_sql_server_utf8(sql)
            for i in range(len(res)):
                item = res[i][0]
                listbox_value1.append(item)
            window1['LIST-BOX'].update(listbox_value1)

        if event == 'LOOK-APPROVE':
            list_user = []
            listbox_select_value = values['LIST-BOX']
            count = 1
            if len(listbox_select_value) != 0:
                approve_name = listbox_select_value[0]
                query_para = 'user1,user2,user3,user4,user5,user6,user7,user8,user9,user10'
                sql = f"SELECT {query_para} FROM approve_user WHERE approve_name=N'{approve_name}' "
                res = exec_sql_server_utf8(sql)
                for item in res[0]:
                    if item not in (None, ''):
                        list_user.append(item)

                if sec_window is not None:
                    sec_window.close()
                window_look_approve = mw.make_window_look_approve()
                sec_window = window_look_approve
                user_number = len(list_user)
                for i in range(user_number):
                    arrow_png = 'ARROW-PNG' + str(1 + i)
                    text_png = 'TEXT-PNG' + str(1 + i)
                    user_item = list_user[i]
                    window_look_approve[f'{text_png}'].update(f'{user_item}')
                    window_look_approve[f'{arrow_png}'].update(filename=pic_arrow)
                    count += 1
                text_end = 'TEXT-PNG' + str(count)
                window_look_approve[f'{text_end}'].update(' 结束 ')
            else:
                sg.popup_auto_close('流程不能为空,请选择流程', title='注意', font=('宋体', 18))

        if event == 'FILE-INPUT' or event == 'ADD-FILE':
            list_id = []
            # list_table = []
            list_filename = []
            approve_table = 'approve_file'
            filepath = values['FILE-INPUT']
            if filepath != '':
                file_number = filepath.split(';')
                approve_file_id = 0
                window1['FILE-INPUT'].update('')
                for item in file_number:
                    file_path = item.replace(r'/', r'\\')
                    addfile_filename = file_path.split('\\')[-1]
                    list_filename.append(addfile_filename)
                    res_file_count = exec_sql_server_utf8(f"SELECT COUNT(1) FROM approve_file "
                                                          f"WHERE file_name=N'{addfile_filename}' AND finish_flag !=2 ")
                    f_count = res_file_count[0][0]
                    print(f_count)
                    if f_count == 0:
                        res_max_id = exec_sql_server_utf8(f"SELECT MAX(ID) FROM {approve_table}")
                        try:
                            approve_file_id = int(res_max_id[0][0]) + 1
                        except Exception as e:
                            approve_file_id = 1
                            print(e)
                        list_id.append(approve_file_id)
                        file_size = get_FileSize(file_path)
                        upload_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                        insert_para = '(ID,file_name,filepath,file_size,upload_time,id_number,finish_flag,finish_status,server_status,popup_flag)'
                        insert_value = (
                            approve_file_id, addfile_filename, item, file_size, upload_time, login_user, 0, '', '', 0)
                        sql1 = f"INSERT INTO {approve_table} {insert_para} VALUES {insert_value}"
                        update_sql_server(sql1)
                        res_copy = copyfile(item, server_temp_path)  # 复制文件
                        if res_copy == 1:
                            logger.info(f'{item}复制成功')
                        else:
                            logger.info(f'{item}复制失败')
                    else:
                        sg.popup_error(f'{addfile_filename}文件名重复，请修改文件名', title='文件名重复', font=('宋体', 18))

                print('list_id列表:', list_id)
                if len(list_id) == 1:
                    item_id = list_id[0]
                    sql2 = f"SELECT {approve_file_para} FROM {approve_table} WHERE ID={item_id}"
                    list_table = get_db_table(sql2)
                    window1['TABLE2'].update(values=list_table)

                elif len(list_id) > 1:
                    item_id = tuple(list_id)
                    sql2 = f"SELECT {approve_file_para} FROM {approve_table} WHERE ID in {item_id}"
                    list_table = get_db_table(sql2)
                    window1['TABLE2'].update(values=list_table)

                else:
                    pass

        if event == 'FILE-INFO':
            seleted_row = values['TABLE2']
            if len(seleted_row) == 0:
                sg.popup_auto_close('行不能为空,请选择对应行', title='注意', font=('宋体', 18))
            else:
                if sec_window is not None:
                    sec_window.close()
                window_file_info = mw.make_window_file_info()
                sec_window = window_file_info
                if len(list_table) > 0:
                    job_id = list_table[seleted_row[0]][2]
                    query_para = 'file_name,Uname,file_size,upload_time,send_time,id_number,finish_status,server_path'
                    sql = f"SELECT {query_para} FROM approve_file WHERE ID = {job_id}"
                    res = exec_sql_server_utf8(sql)
                    window_file_info['FILE-INFO-ID'].update(job_id)
                    window_file_info['FILE-INFO-NAME'].update(res[0][0])
                    window_file_info['FILE-INFO-APPROVE'].update(res[0][1])
                    window_file_info['FILE-INFO-FILESIZE'].update(res[0][2])
                    window_file_info['FILE-INFO-UPLOAD-TIME'].update(res[0][3])
                    window_file_info['FILE-INFO-SEND-TIME'].update(res[0][4])
                    window_file_info['FILE-INFO-ID-NUMBER'].update(res[0][5])
                    window_file_info['FILE-INFO-FILE-STATUS'].update(res[0][6])
                    window_file_info['FILE-INFO-SERVER-PATH'].update(res[0][7])

        if event == 'ADD-APPROVE':
            if sec_window is not None:
                sec_window.close()
            window_add_approve = mw.make_window_add_approve()
            sec_window = window_add_approve
            id_value = table_max_id('approve_user')
            window_add_approve['ADD-ID'].update(id_value)
            update_combo_user1(window_add_approve)

        if event == 'ADD-APPROVE-SAVE':
            add_type_value = 1
            add_id = values['ADD-ID']
            add_type = values['ADD-TYPE']
            add_name = values['ADD-NAME']
            add_user1 = values['ADD-USER1']
            add_user2 = values['ADD-USER2']
            add_user3 = values['ADD-USER3']
            add_user4 = values['ADD-USER4']
            add_user5 = values['ADD-USER5']
            add_user6 = values['ADD-USER6']
            add_user7 = values['ADD-USER7']
            add_user8 = values['ADD-USER8']
            add_user9 = values['ADD-USER9']
            add_user10 = values['ADD-USER10']
            if add_name == '':
                sg.popup_auto_close('流程名称不能为空', title='注意', font=('宋体', 18))
            else:
                res_name = exec_sql_server_utf8(f"SELECT ID FROM approve_user WHERE approve_name=N'{add_name}' ")
                if len(res_name) == 0:
                    if add_type == '临技下发':
                        add_type_value = 2

                    insert_para = '(ID,approve_name,user1,user2 ,user3,user4 ,user5,' \
                                  'user6,user7,user8,user9,user10,file_type)'
                    insert_value = (add_id, add_name, add_user1, add_user2, add_user3, add_user4, add_user5, add_user6,
                                    add_user7, add_user8, add_user9, add_user10, add_type_value)
                    res = update_sql_server(f"INSERT INTO approve_user {insert_para} VALUES {insert_value}")

                    if add_type_value == 1:
                        file_review_flag = 1
                    else:
                        temp_release_flag = 1

                    if res == 1:

                        logger.info(f"流程{add_name}保存成功")
                        window_add_approve.close()
                        sg.popup_auto_close('保存成功', title='成功', font=('宋体', 18))
                    else:
                        logger.info(f"流程{add_name}保存失败")
                        sg.popup_auto_close('保存失败', title='失败', font=('宋体', 18))
                else:
                    sg.popup_auto_close('保存失败,流程名称不能相同', title='失败', font=('宋体', 18))

        if event == 'DEL-APPROVE':
            listbox_select_value = values['LIST-BOX']
            if len(listbox_select_value) != 0:
                approve_name = listbox_select_value[0]
                sql = f"DELETE approve_user WHERE approve_name=N'{approve_name}' "
                sql1 = f"SELECT file_type from approve_user WHERE approve_name=N'{approve_name}'"
                res = update_sql_server(sql)
                if res == 1:
                    logger.info(f"{approve_name}删除成功")
                    sg.popup_auto_close(f"{approve_name}删除成功", title='成功', font=('宋体', 18))
                    file_review_flag = 1
                else:
                    logger.info(f"{approve_name}删除失败")
                    sg.popup_auto_close(f"{approve_name}删除失败", title='失败', font=('宋体', 18))

            else:
                sg.popup_auto_close('流程不能为空,请选择流程', title='注意', font=('宋体', 18))

        if event == 'EDIT-APPROVE':
            query_para = 'user1,user2,user3,user4,user5,user6,user7,user8,user9,user10,file_type,ID'
            listbox_select_value = values['LIST-BOX']
            if len(listbox_select_value) != 0:
                approve_name = listbox_select_value[0]
                sql = f"SELECT {query_para} FROM approve_user WHERE approve_name=N'{approve_name}' "
                res = exec_sql_server_utf8(sql)
                if sec_window is not None:
                    sec_window.close()
                window_edit_approve = mw.make_window_edit_approve()
                sec_window = window_edit_approve
                window_edit_approve['EDIT-ID'].update(res[0][-1])
                window_edit_approve['EDIT-NAME'].update(approve_name)
                update_combo_user2(window_edit_approve)
                for i in range(10):
                    item_value = res[0][i]
                    item_key = 'EDIT-USER' + str(i + 1)
                    window_edit_approve[f'{item_key}'].update(item_value)
                file_type_value = res[0][-2]
                if file_type_value == 1:
                    window_edit_approve['EDIT-TYPE'].update('文件审批')
                else:
                    window_edit_approve['EDIT-TYPE'].update('临技下发')

            else:
                sg.popup_auto_close('流程不能为空,请选择流程', title='注意', font=('宋体', 18))

        if event == 'EDIT-APPROVE-SAVE':
            edit_file_type_value = 1
            edit_id = values['EDIT-ID']
            edit_type = values['EDIT-TYPE']
            edit_name = values['EDIT-NAME']
            edit_user1 = values['EDIT-USER1']
            edit_user2 = values['EDIT-USER2']
            edit_user3 = values['EDIT-USER3']
            edit_user4 = values['EDIT-USER4']
            edit_user5 = values['EDIT-USER5']
            edit_user6 = values['EDIT-USER6']
            edit_user7 = values['EDIT-USER7']
            edit_user8 = values['EDIT-USER8']
            edit_user9 = values['EDIT-USER9']
            edit_user10 = values['EDIT-USER10']
            edit_file_type = values['EDIT-TYPE']
            if edit_file_type == '临技下发':
                edit_file_type_value = 2
            res = update_sql_server(
                f"UPDATE approve_user set file_type={edit_file_type_value}, approve_name='{edit_name}',"
                f"user1='{edit_user1}', user2='{edit_user2}', user3='{edit_user3}', user4='{edit_user4}', "
                f"user5='{edit_user5}', user6='{edit_user6}', user7='{edit_user7}', user8='{edit_user8}', "
                f"user9='{edit_user9}', user10='{edit_user10}' "
                f"WHERE ID={edit_id} ")
            if res == 1:
                logger.info(f"流程{edit_name}修改成功")
                window_edit_approve.close()
                # file_review_flag = 1
                sg.popup_auto_close('修改成功', title='成功', font=('宋体', 18))
            else:
                logger.info(f"流程{edit_name}修改失败")
                sg.popup_auto_close('修改失败', title='失败', font=('宋体', 18))

        if event == 'SEND-APPROVE':
            error_flag = 0  # 错误标志
            list_user = []
            list_user_id = []
            file_type_value = 1
            listbox_value = values['LIST-BOX']
            user_content = values['MT-OUTPUT'].strip()  # 去除首尾空格
            if len(list_id) != 0:
                if len(listbox_value) != 0:
                    for id in list_id:
                        list_user = []
                        list_user_id = []
                        approve_name = listbox_value[0]
                        user_action = '发送申请'
                        query_para = 'user1,user2,user3,user4,user5,user6,user7,user8,user9,user10'
                        insert_para = '(job_id,id_number,user_action,user_content,enter_time,finish_flag,' \
                                      'approve_level,job_flag)'
                        send_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                        sql1 = f"SELECT {query_para} FROM approve_user WHERE approve_name=N'{approve_name}' "
                        res = exec_sql_server_utf8(sql1)
                        for item in res[0]:
                            if item not in (None, ''):
                                list_user.append(item)
                        for item in list_user:  # 通过用户名查询用户ID
                            res_id = exec_sql_server_utf8(f"SELECT id_number FROM member WHERE name=N'{item}' ")
                            if len(res_id) > 0:
                                list_user_id.append(res_id[0][0])
                            else:
                                sg.popup_error(f'未找到用户:{item}, 请检查审批流程：{approve_name} 该用户名称是否正确!', title='错误',
                                               font=('宋体', 18))
                                error_flag = 1
                                break
                        if error_flag == 1:
                            break  # 退出循环
                        else:
                            if len(list_user) > 0:
                                wait_user = list_user[0]
                            print(list_user, list_user_id)
                            print('待审批人:', wait_user)
                            # approve_file更新流程名称,id
                            update_sql_server(f"UPDATE approve_file set Uname=N'{approve_name}',send_time='{send_time}', "
                                              f"id_number='{login_user}', finish_status=N'待{wait_user}审批' "
                                              f"WHERE ID={id}")
                            sql2 = f"SELECT file_type FROM approve_user WHERE approve_name=N'{approve_name}' "
                            file_type = exec_sql_server_utf8(sql2)
                            file_type_value = file_type[0][0]
                            if file_type_value == 1:  # approve_job 写入内容，1为文件审批流程
                                insert_value1 = (id, login_user, user_action, user_content, send_time, '1', 0, 2)
                                update_sql_server(f"INSERT INTO approve_job {insert_para} VALUES {insert_value1}")

                                for i in range(len(list_user)):
                                    id_num = list_user_id[i]
                                    insert_value2 = (id, id_num, '', '', '', '', i + 1, 1)
                                    update_sql_server(f"INSERT INTO approve_job {insert_para} VALUES {insert_value2}")

                            else:
                                insert_value1 = (id, login_user, user_action, user_content, send_time, '1', 0, 2)
                                update_sql_server(f"INSERT INTO approve_job {insert_para} VALUES {insert_value1}")

                                for i in range(len(list_user)):
                                    id_num = list_user_id[i]
                                    insert_value2 = (id, id_num, '', '', '', '', 1, 1)
                                    update_sql_server(f"INSERT INTO approve_job {insert_para} VALUES {insert_value2}")
                    if error_flag == 0:

                        # 更新页面显示
                        if len(list_id) == 1:
                            sql = f"SELECT {approve_file_para} FROM approve_file WHERE ID={list_id[0]} "
                        else:
                            tuple_list_id = tuple(list_id)
                            sql = f"SELECT {approve_file_para} FROM approve_file WHERE ID in {tuple_list_id}"
                        list_table = get_db_table(sql)
                        window1['TABLE2'].update(values=list_table)

                        # 写入lj_document表
                        if file_type_value == 2:
                            table_name = 'lj_document'
                            for i in range(len(list_filename)):
                                id_number = table_max_id(table_name)
                                file_code = list_filename[i].split('.')
                                print(file_code)

                                if file_code[0][0:4] == 'L202' or file_code[0][0:4] == 'C202':
                                    now_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                                    insert_item = '(ID,file_code,product_code,rwl,createdate,validdate,station,lj_type,remark,file_path,job_id)'
                                    insert_value = (id_number, file_code[0], '', '', now_time, '', '', '', '', '', id)
                                    # 写入lj_document表
                                    update_sql_server(f"INSERT INTO {table_name} {insert_item} VALUES {insert_value} ")
                                    # 文件复制到指定目录
                                    lj_path = r'\\10.10.80.80\share\宇视\临技\已处理'
                                    lj_temp_path = server_temp_path + '\\' + addfile_filename
                                    copyfile(lj_temp_path, lj_path)
                                else:
                                    pass
                        list_id = []  # 清空
                        list_filename = []
                        sg.popup_auto_close('发送成功', title='成功', font=('宋体', 18))

                else:
                    sg.popup_auto_close('流程不能为空，请选择流程', title='注意', font=('宋体', 18))

            else:
                sg.popup_auto_close('文件不能为空，请添加文件', title='注意', font=('宋体', 18))

        if event == 'OPEN-FILE':
            seleted_row = values['TABLE2']

            if len(seleted_row) == 0:
                sg.popup_auto_close('行不能为空,请选择对应行', title='注意', font=('宋体', 18))
            else:
                if len(list_table) > 0:
                    job_id = list_table[seleted_row[0]][2]
                    doc_name = exec_sql_server_utf8(f"SELECT file_name from approve_file WHERE ID={job_id}")
                    doc_path = server_temp_path + '\\' + doc_name[0][0]
                    try:
                        os.startfile(doc_path)
                    except Exception as e:
                        sg.popup_error(e, font=('宋体', 18))

        if event == 'DEL-FILE':
            list_id = []
            seleted_row = values['TABLE2']
            if len(seleted_row) == 0:
                sg.popup_auto_close('行不能为空,请选择对应行', title='注意', font=('宋体', 18))
            else:
                if len(list_table) > 0:
                    for id in seleted_row:
                        job_id = list_table[id][2]
                        list_id.append(job_id)
                    if sec_window is not None:
                        sec_window.close()
                    window_make_sure = mw.makesure_window()
                    sec_window = window_make_sure

        if window == window_make_sure and event == 'ENTER-BOTTON':
            count = 0
            id_value = ''
            for id in list_id:
                res_id = exec_sql_server_utf8(f"SELECT id_number FROM approve_file WHERE ID={id} ")
                try:
                    id_value = res_id[0][0].replace(' ', '')  # 去掉空格
                except Exception as e:
                    sg.popup_error(e, title='错误', font=('宋体', 18))
                if id_value == login_user:
                    res1 = update_sql_server(f"DELETE approve_file  WHERE ID={id} ")
                    res2 = update_sql_server(f"DELETE approve_job WHERE job_id={id} ")
                    my_send_flag = 1
                    window_make_sure.close()
                    count = count + res1 + res2
                elif login_user == 'root':
                    res1 = update_sql_server(f"DELETE approve_file  WHERE ID={id} ")
                    res2 = update_sql_server(f"DELETE approve_job WHERE job_id={id} ")
                    my_send_flag = 1
                    window_make_sure.close()
                    count = count + res1 + res2
                else:
                    window_make_sure.close()
                    sg.popup_auto_close('删除失败，需申请人删除该文件', title='失败', font=('宋体', 18))
            if len(list_id) * 2 == count:
                logger.info(f"任务ID:{list_id}删除成功")
                sg.popup_auto_close(f'任务ID:{list_id}删除成功', title='成功', font=('宋体', 18))
            else:
                logger.info(f"任务ID:{list_id}删除失败")
                sg.popup_auto_close(f'任务ID:{list_id}删除失败', title='失败', font=('宋体', 18))

        if event == 'DOWN-FILE':
            seleted_row = values['TABLE2']
            print(seleted_row)
            if len(seleted_row) == 0:
                sg.popup_auto_close('行不能为空,请选择对应行', title='注意', font=('宋体', 18))
            else:
                if len(list_table) > 0:
                    for item_num in seleted_row:
                        job_id = list_table[item_num][2]
                        res = exec_sql_server_utf8(f"SELECT file_name,server_path FROM approve_file WHERE ID={job_id} ")
                        doc_name = res[0][0]
                        doc_server_path = res[0][1]
                        if doc_server_path is not None:
                            res1 = copyfile(doc_server_path, download_path)
                            if res1 == 1:
                                logger.info(f'{doc_name} 下载成功,下载路径:{download_path}')
                                sg.popup_auto_close(f'{doc_name} 下载成功,下载路径:{download_path}', title='成功',
                                                    font=('宋体', 18))
                            else:
                                sg.popup_auto_close(f'{doc_name} 下载失败', title='失败', font=('宋体', 18))
                        else:
                            new_path = server_temp_path + '\\' + doc_name
                            res1 = copyfile(new_path, download_path)
                            if res1 == 1:
                                logger.info(f'{doc_name} 下载成功,下载路径:{download_path}')
                                sg.popup_auto_close(f'{doc_name} 下载成功,下载路径:{download_path}', title='成功',
                                                    font=('宋体', 18))
                            else:
                                sg.popup_auto_close(f'{doc_name} 下载失败', title='失败', font=('宋体', 18))

        if event == 'APPROVE-INFO' or event == 'TABLE2Double':
            seleted_row = values['TABLE2']

            if len(seleted_row) == 0:
                sg.popup_auto_close('行不能为空,请选择对应行', title='注意', font=('宋体', 18))
            else:
                if sec_window is not None:
                    sec_window.close()
                window_approve_info = mw.make_winodw_approve_info()
                sec_window = window_approve_info
                window_approve_info['TABLE-APP-INFO'].bind("<Double-Button-1>", 'Double')
                if login_user == 'root':
                    window_approve_info['APPROVE-RETRACT'].update(visible=True)

                if len(list_table) > 0:
                    print(list_table[seleted_row[0]])
                    job_id = list_table[seleted_row[0]][2]
                    job_name = list_table[seleted_row[0]][1]
                    job_name_file = job_name.split('/')
                    print('任务ID:', job_id)
                    uname = exec_sql_server_utf8(f"SELECT Uname FROM approve_file WHERE ID='{job_id}' ")
                    sql = f"SELECT b.name,a.user_action,a.user_content,a.enter_time " \
                          f"FROM approve_job AS a JOIN member AS b ON a.id_number = b.id_number " \
                          f"where a.job_id={job_id} ORDER BY approve_level"
                    list_table3 = get_db_table(sql)
                    window_approve_info['APPROVE-TEXT-INFO'].update(job_name_file[-1])
                    window_approve_info['APPROVE-NAME-INFO'].update(uname[0][0])
                    window_approve_info['TABLE-APP-INFO'].update(values=list_table3)

        if event == 'TABLE-APP-INFODouble':
            row_number = values['TABLE-APP-INFO']
            row_info = list_table3[row_number[0]]
            user_content = row_info[2]
            window_display_user_content = mw.make_window_display_user_content()
            window_display_user_content['USER-CONTENT-OUTPUT'].update(user_content)

        if event == 'ONLINE-USER':
            if sec_window is not None:
                sec_window.close()
            window_online_user = mw.make_winodw_online_user()
            sec_window = window_online_user
            sql = f"SELECT ROW_NUMBER() OVER ( ORDER BY user_status DESC ) AS row_number, id_number,name,user_status FROM member " \
                  f"WHERE user_status in (0,1) and role in (0,1)"
            list_table_user = get_db_table(sql)
            window_online_user['TABLE-ONLINE-USER'].update(values=list_table_user)
            window_online_user['TABLE-ONLINE-USER'].bind("<Double-Button-1>", 'Double')

        if event == 'EDIT-USER-INFO' or event == 'TABLE-ONLINE-USERDouble':
            seleted_row = values['TABLE-ONLINE-USER']
            row_number = seleted_row[0]
            id = list_table_user[row_number][1]
            window_edit_user_info = mw.make_window_edit_user_info()
            query_para = 'id_password,name,department,position,role,user_status'
            res = exec_sql_server_utf8(f"SELECT {query_para} FROM member WHERE id_number='{id}' ")
            paswd = res[0][0]
            name = res[0][1]
            depart = res[0][2]
            position = res[0][3]
            role_value = res[0][4]
            user_status = res[0][5]
            if role_value == 1:
                role = '管理员'
            else:
                role = '用户'
            if user_status == 1:
                user_status_cn = '在线'
            else:
                user_status_cn = '不在线'
            window_edit_user_info['DIS-LOGIN-USER-INFO'].update(id)
            window_edit_user_info['DIS-LOGIN-PASSWD-INFO'].update(paswd)
            window_edit_user_info['DIS-USERNAME-INFO'].update(name)
            window_edit_user_info['DIS-DEPART-INFO'].update(depart)
            window_edit_user_info['DIS-POSITION-INFO'].update(position)
            window_edit_user_info['DIS-ROLE-INFO'].update(role)
            window_edit_user_info['DIS-USER-INFO'].update(user_status_cn)

        if event == 'DEL-USER-INFO':
            seleted_row = values['TABLE-ONLINE-USER']
            row_number = seleted_row[0]
            print(list_table_user[row_number])
            id = list_table_user[row_number][1]
            res = update_sql_server(f"DELETE member WHERE id_number='{id}' ")
            sql = f"SELECT ROW_NUMBER() OVER ( ORDER BY user_status DESC ) AS row_number, id_number,name,user_status FROM member " \
                  f"WHERE user_status in (0,1) "
            list_table_user = get_db_table(sql)
            window_online_user['TABLE-ONLINE-USER'].update(values=list_table_user)
            if res == 1:
                logger.info(f'用户{id} 删除成功')
                sg.popup_auto_close('删除成功', title='成功', font=('宋体', 18))
            else:
                logger.info(f'用户{id} 删除失败')
                sg.popup_auto_close('删除失败', title='失败', font=('宋体', 18))

        if event == 'EDIT-USER-SAVE-INFO':
            user_id = values['DIS-LOGIN-USER-INFO']
            login_passwd = values['DIS-LOGIN-PASSWD-INFO']
            username = values['DIS-USERNAME-INFO']
            depart = values['DIS-DEPART-INFO']
            potision = values['DIS-POSITION-INFO']
            role = values['DIS-ROLE-INFO']
            user_status_cn = values['DIS-ROLE-INFO']
            res_passwd = exec_sql_server_utf8(f"SELECT id_password FROM member WHERE id_number='{user_id}' ")
            print(login_passwd, res_passwd[0][0])
            if role == '管理员':
                role_value = 1
            else:
                role_value = 0

            if user_status_cn == '在线':
                user_status = 1
            else:
                user_status = 0

            if user_id == '':
                sg.popup_auto_close('账号不能为空', title='注意', font=('宋体', 18))
            elif login_passwd == '':
                sg.popup_auto_close('密码不能为空', title='注意', font=('宋体', 18))
            elif username == '':
                sg.popup_auto_close('姓名不能为空', title='注意', font=('宋体', 18))
            else:
                if login_passwd == res_passwd[0][0]:
                    res = update_sql_server(f"UPDATE member SET name=N'{username}',department=N'{depart}',"
                                            f"position='{potision}',role={role_value} ,user_status={user_status} "
                                            f"WHERE id_number='{user_id}' ")
                else:
                    login_passwd_md5 = computerMD5(login_passwd)
                    res = update_sql_server(f"UPDATE member SET id_password='{login_passwd_md5}',"
                                            f"name=N'{username}',department=N'{depart}',"
                                            f"position='{potision}',role={role_value} ,user_status={user_status} "
                                            f"WHERE id_number='{user_id}' ")
                if res == 1:
                    window_edit_user_info.close()
                    logger.info(f'修改用户{username}信息成功')
                    sg.popup_auto_close('修改用户信息成功', title='成功', font=('宋体', 18))
                else:
                    logger.info(f'修改用户{username}信息失败')
                    sg.popup_auto_close('修改用户信息失败', title='失败', font=('宋体', 18))

        if event == 'ADD-USER':
            if sec_window is not None:
                sec_window.close()
            window_add_user = mw.make_window_add_user()
            sec_window = window_add_user

        if event == 'ADD-USER-SAVE':
            add_login_user = values['ADD-LOGIN-USER']
            add_login_passwd = values['ADD-LOGIN-PASSWD']
            add_username = values['ADD-USERNAME']
            add_depart = values['ADD-DEPART']
            add_potision = values['ADD-POSITION']
            add_role = values['ADD-ROLE']
            add_login_passwd_md5 = computerMD5(add_login_passwd)
            if add_role == '管理员':
                add_role_value = 1
            else:
                add_role_value = 0
            if add_login_user == '':
                sg.popup_auto_close('账号不能为空', title='注意', font=('宋体', 18))
            elif add_login_passwd == '':
                sg.popup_auto_close('密码不能为空', title='注意', font=('宋体', 18))
            elif add_username == '':
                sg.popup_auto_close('姓名不能为空', title='注意', font=('宋体', 18))
            else:
                res1 = exec_sql_server_utf8(f"SELECT name FROM member WHERE id_number='{add_login_user}' ")
                if len(res1) != 0:
                    sg.popup_auto_close('账号重复', title='注意', font=('宋体', 18))
                else:
                    insert_para = '(id_number,id_password,name,department,position,role,user_status)'
                    insert_value = (add_login_user, add_login_passwd_md5, add_username, add_depart,
                                    add_potision, add_role_value, 0)
                    res2 = update_sql_server(f"INSERT INTO member {insert_para} VALUES {insert_value}")
                    if res2 == 1:
                        window_add_user.close()
                        logger.info(f'新增用户{add_login_user}成功')
                        sg.popup_auto_close('新增用户成功', title='成功', font=('宋体', 18))
                    else:
                        logger.info(f'新增用户{add_login_user}失败')
                        sg.popup_auto_close('新增用户失败', title='失败', font=('宋体', 18))

        if event == 'USER-PNG':
            if sec_window is not None:
                sec_window.close()
            window_edit_user = mw.make_window_edit_user()
            sec_window = window_edit_user
            query_para = 'id_password,name,department,position,role'
            res = exec_sql_server_utf8(f"SELECT {query_para} FROM member WHERE id_number='{login_user}' ")
            paswd = res[0][0]
            name = res[0][1]
            depart = res[0][2]
            position = res[0][3]
            role_value = res[0][4]
            if role_value == 1:
                role = '管理员'
            else:
                role = '用户'
            window_edit_user['EDIT-LOGIN-USER'].update(login_user)
            window_edit_user['EDIT-LOGIN-PASSWD'].update(paswd)
            window_edit_user['EDIT-USERNAME'].update(name)
            window_edit_user['EDIT-DEPART'].update(depart)
            window_edit_user['EDIT-POSITION'].update(position)
            window_edit_user['EDIT-ROLE'].update(role)

        if event == 'EDIT-USER-SAVE':
            edit_login_user = values['EDIT-LOGIN-USER']
            edit_login_passwd = values['EDIT-LOGIN-PASSWD']
            edit_username = values['EDIT-USERNAME']
            edit_depart = values['EDIT-DEPART']
            edit_potision = values['EDIT-POSITION']
            edit_role = values['EDIT-ROLE']
            res_passwd = exec_sql_server_utf8(f"SELECT id_password FROM member WHERE id_number='{edit_login_user}' ")
            print(edit_login_passwd, res_passwd[0][0])
            if edit_role == '管理员':
                edit_role_value = 1
            else:
                edit_role_value = 0
            if edit_login_passwd == '':
                sg.popup_auto_close('密码不能为空', title='注意', font=('宋体', 18))
            elif edit_username == '':
                sg.popup_auto_close('姓名不能为空', title='注意', font=('宋体', 18))
            else:
                if edit_login_passwd == res_passwd[0][0]:
                    res = update_sql_server(f"UPDATE member SET name=N'{edit_username}',department='{edit_depart}',"
                                            f"position='{edit_potision}',role={edit_role_value} "
                                            f"WHERE id_number='{login_user}' ")
                else:
                    edit_login_passwd_md5 = computerMD5(edit_login_passwd)
                    res = update_sql_server(f"UPDATE member SET id_password='{edit_login_passwd_md5}',"
                                            f"name=N'{edit_username}',department='{edit_depart}',"
                                            f"position='{edit_potision}',role={edit_role_value} "
                                            f"WHERE id_number='{login_user}' ")
                if res == 1:
                    window_edit_user.close()
                    logger.info(f'修改用户{login_user}信息成功')
                    sg.popup_auto_close('保存成功', title='成功', font=('宋体', 18))
                else:
                    logger.info(f'修改用户{login_user}信息失败')
                    sg.popup_auto_close('保存失败', title='失败', font=('宋体', 18))

        if event == 'MY-SEND' or my_send_flag == 1:
            my_send_flag = 0
            approve_status = window1['APP-STATUS-INPUT'].get()
            frame4_input = window1['FRAME4-INPUT'].get()
            user_id = window1['ID-NUMBER-INPUT'].get()
            date_input = window1['DATE-INPUT'].get()
            if approve_status == '审批中':
                approve_status = '审批'
            if login_user == 'root':
                sql = f"SELECT {approve_file_para} FROM approve_file " \
                      f"WHERE file_name like '%{frame4_input}%' and id_number like '%{user_id}%' and finish_status LIKE '%{approve_status}%' " \
                      f"and send_time >= '{date_input}' "
            else:
                sql = f"SELECT {approve_file_para} FROM approve_file WHERE id_number='{login_user}' " \
                      f"and file_name like '%{frame4_input}%' and finish_status LIKE '%{approve_status}%' " \
                      f"and send_time >= '{date_input}' "
            list_table = get_db_table(sql)
            window1['TABLE2'].update(values=list_table)

        if event == 'WAIT-APPROVE' or wait_approve_flag == 1:
            print('job标志位:', check_job_flag)
            wait_approve_flag = 0
            frame4_input = window1['FRAME4-INPUT'].get()
            user_id = window1['ID-NUMBER-INPUT'].get()
            date_input = window1['DATE-INPUT'].get()

            if login_user == 'root':
                sql = f"SELECT {approve_file_para} FROM approve_file " \
                      f"WHERE ID IN (SELECT job_id FROM approve_job WHERE job_flag=1) " \
                      f"and file_name like '%{frame4_input}%' and id_number like '%{user_id}%' " \
                      f"and send_time >= '{date_input}'"
            else:
                sql = f"SELECT {approve_file_para} FROM approve_file " \
                      f"WHERE ID IN (SELECT job_id FROM approve_job WHERE job_flag=1 and id_number='{login_user}') " \
                      f"and file_name like '%{frame4_input}%' and id_number like '%{user_id}%' " \
                      f"and send_time >= '{date_input}'"
            list_table_wait_approve = get_db_table(sql)
            list_table = list_table_wait_approve
            window1['TABLE2'].update(values=list_table)

        if event == 'FINISH-APPROVE' or finish_approve_flag == 1:
            finish_approve_flag = 0
            list_table = []
            frame4_input = window1['FRAME4-INPUT'].get()
            user_id = window1['ID-NUMBER-INPUT'].get()
            approve_status = window1['APP-STATUS-INPUT'].get()
            date_input = window1['DATE-INPUT'].get()
            if approve_status == '审批中':
                approve_status = '审批'

            approve_file_para1 = 'ROW_NUMBER() OVER ( ORDER BY ID ) AS row_number,' \
                                 'file_name,a.ID,a.id_number,a.send_time,a.finish_status,a.Uname'
            if login_user == 'root':
                sql = f"SELECT {approve_file_para} FROM approve_file " \
                      f"WHERE finish_flag=1 and file_name like '%{frame4_input}%' and id_number like '%{user_id}%' " \
                      f"and finish_status LIKE '%{approve_status}%' and send_time >= '{date_input}' "
            else:
                sql = f"SELECT {approve_file_para1} FROM approve_file AS a JOIN approve_job AS b on a.ID=b.job_id " \
                      f"WHERE b.finish_flag=1 and b.approve_level <> 0 and b.id_number='{login_user}' " \
                      f"and a.file_name like '%{frame4_input}%' and a.finish_status LIKE '%{approve_status}%' " \
                      f"and a.id_number LIKE '%{user_id}%' and a.send_time >= '{date_input}' "
            list_table = get_db_table(sql)
            window1['TABLE2'].update(values=list_table)

        if event == 'REJECT-SEND':
            list_table = []
            frame4_input = window1['FRAME4-INPUT'].get()
            user_id = window1['ID-NUMBER-INPUT'].get()
            date_input = window1['DATE-INPUT'].get()
            if login_user == 'root':
                sql = f"SELECT {approve_file_para} FROM approve_file " \
                      f"WHERE finish_flag=2 and file_name like '%{frame4_input}%' and id_number like '%{user_id}%' " \
                      f"and send_time >= '{date_input}'"
            else:
                sql = f"SELECT {approve_file_para} FROM approve_file " \
                      f"WHERE finish_flag=2 and id_number='{login_user}' and file_name like '%{frame4_input}%' " \
                      f"and send_time >= '{date_input}'"
            list_table = get_db_table(sql)
            window1['TABLE2'].update(values=list_table)

        if event == 'APPROVE-EXPORT':
            print(list_table)
            now_time = time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime(time.time()))
            new_sp_excel = excel_path + '\\' + 'SP' + f'-{now_time}.xlsx'
            wb = Workbook()  # 创建工作薄对象
            ws = wb.create_sheet('审批', 0)  # 创建工作表对象,0对应第1个sheet，1为第2个sheet
            if len(list_table) != 0:
                try:
                    for i in range(len(list_table)):
                        item = list_table[i]
                        print(item)
                        write_row_by_list(ws, i + 1, item)
                    wb.save(new_sp_excel)
                    logger.info(f'导出成功,路径:{new_sp_excel}')
                    sg.popup_ok(f'路径:{new_sp_excel}', title='导出成功', font=('宋体', 18))

                except Exception as result:
                    sg.popup_error(result, font=('宋体', 18))
            else:
                sg.popup_auto_close('请先查询，再导出', title='注意', font=('宋体', 18))

        if event == 'FILE-APPROVE':
            png_file_path = r'\\10.10.80.80\share\临时共享\temp_document\png\file_approve.png'
            try:
                os.startfile(png_file_path)
            except Exception as e:
                sg.popup_error(e, title='错误', font=('宋体', 18))

        if event == 'AGREE-APPROVE':
            count = 0
            multi_input = values['MT-OUTPUT'].strip()
            seleted_row = values['TABLE2']
            if len(seleted_row) == 0:
                sg.popup_auto_close('文件不能为空请选择文件', title='注意', font=('宋体', 18))
            else:
                if len(list_table_wait_approve) == 0:
                    sg.popup_auto_close('没有待审批文件,请选择待审批文件', title='注意', font=('宋体', 18))
                else:
                    for row in seleted_row:
                        job_id = list_table[row][2]
                        file_name = list_table[row][1]
                        print('job_id , file_name', job_id, file_name)
                        user_action = '同意'
                        enter_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))

                        # 更新approve_job,写入相关记录
                        sql = f"SELECT b.file_type FROM approve_file AS a " \
                              f"JOIN approve_user AS b on a.Uname = b.approve_name " \
                              f"WHERE a.ID={job_id} "
                        file_type = exec_sql_server_utf8(sql)
                        if len(file_type) != 0:
                            file_type_value = file_type[0][0]
                            if file_type_value == 1:
                                id_number = exec_sql_server_utf8(f"SELECT id_number FROM approve_job "
                                                                 f"WHERE job_id={job_id} and finish_flag='0' "
                                                                 f"ORDER BY approve_level ")
                                print('当前审批人ID:', id_number[0][0])
                                if id_number[0][0] == login_user:  # 文件审批判断是否是当前审批人
                                    sql = f"UPDATE approve_job SET user_action=N'{user_action}',user_content=N'{multi_input}', " \
                                          f"enter_time='{enter_time}',finish_flag=1,job_flag=2 " \
                                          f"WHERE job_id={job_id} and id_number='{login_user}' and approve_level <> 0"
                                    res = update_sql_server(sql)
                                    count = count + res
                                    print('审批数量:', count)

                                else:
                                    sg.popup_auto_close('审批失败,审批人与当前用户不一致', title='审批失败', font=('宋体', 18))

                            if file_type_value == 2:
                                sql = f"UPDATE approve_job SET user_action=N'{user_action}',user_content=N'{multi_input}', " \
                                      f"enter_time='{enter_time}',finish_flag=1,job_flag=2 " \
                                      f"WHERE job_id={job_id} and id_number='{login_user}' and approve_level <> 0"
                                res = update_sql_server(sql)
                                count = count + res
                            # 更新审批状态
                            write_finish_flag = exec_sql_server_utf8(f"SELECT id_number from approve_job "
                                                                     f"WHERE job_id={job_id} and finish_flag = 0 ORDER BY approve_level")
                            if len(write_finish_flag) > 0:  # 审批未完成
                                wait_user = write_finish_flag[0][0]
                                user_name = exec_sql_server_utf8(
                                    f"SELECT name from member where id_number='{wait_user}' ")
                                update_sql_server(f"UPDATE approve_file SET finish_status=N'待{user_name[0][0]}审批' "
                                                  f"WHERE ID={job_id} ")

                        # 文件自动归档,更新审批状态
                        write_finish_flag = exec_sql_server_utf8(f"SELECT id_number from approve_job "
                                                                 f"WHERE job_id={job_id} and finish_flag = 0 ORDER BY approve_level")
                        if len(write_finish_flag) == 0:  # 审批已完成
                            update_sql_server(
                                f"UPDATE approve_file SET finish_flag=1,finish_status=N'已完成',popup_flag=1 "
                                f"WHERE ID='{job_id}' ")
                            pop_times = 1  # 写入完成，开启任务弹屏
                            temp_file_name = server_temp_path + '\\' + file_name
                            upload_res = upload_to_server(temp_file_name, file_name)

                            write_db_from_file(file_name)  # 更新对应文件 文件名，版本，时间

                            if upload_res == 0:
                                sg.popup_auto_close('文件归档失败', title='注意', font=('宋体', 18))
                            else:
                                update_sql_server(f"UPDATE approve_file SET server_path=N'{upload_res}',"
                                                  f"finish_status=N'已归档' WHERE ID='{job_id}' ")

                    if len(seleted_row) == count:
                        sg.popup_auto_close('审批成功', title='成功', font=('宋体', 18))
                    else:
                        sg.popup_auto_close('审批失败', title='失败', font=('宋体', 18))

                    # 刷新审批页面
                    # wait_approve_flag = 1
                    sql = f"SELECT {approve_file_para} FROM approve_file " \
                          f"WHERE ID IN (SELECT job_id FROM approve_job WHERE job_flag=1 and id_number='{login_user}') "
                    list_table_wait_approve = get_db_table(sql)
                    list_table = list_table_wait_approve
                    window1['TABLE2'].update(values=list_table)

        if event == 'REJECT-APPROVE':
            count = 0
            seleted_row = values['TABLE2']
            multi_input = values['MT-OUTPUT'].strip()
            if len(seleted_row) == 0:
                sg.popup_auto_close('文件不能为空请选择文件', title='注意', font=('宋体', 18))
            else:
                if len(list_table_wait_approve) == 0:
                    sg.popup_auto_close('没有待审批文件,请选择待审批文件', title='注意', font=('宋体', 18))
                else:
                    for row in seleted_row:
                        job_id = list_table[row][2]
                        user_action = '拒绝'
                        enter_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))

                        sql = f"SELECT b.file_type FROM approve_file AS a " \
                              f"JOIN approve_user AS b on a.Uname = b.approve_name " \
                              f"WHERE a.ID={job_id} "
                        file_type = exec_sql_server_utf8(sql)
                        if len(file_type) != 0:
                            file_type_value = file_type[0][0]
                            if file_type_value == 1:
                                id_number = exec_sql_server_utf8(f"SELECT id_number FROM approve_job "
                                                                 f"WHERE job_id={job_id} and finish_flag='0' "
                                                                 f"ORDER BY approve_level ")
                                print('当前审批人ID:', id_number[0][0])
                                if id_number[0][0] == login_user:  # 文件审批判断是否是当前审批人
                                    sql = f"UPDATE approve_job SET user_action=N'{user_action}',user_content=N'{multi_input}', " \
                                          f"enter_time='{enter_time}',finish_flag=1,job_flag=2 " \
                                          f"WHERE job_id={job_id} and id_number='{login_user}' "
                                    res1 = update_sql_server(sql)
                                    res2 = update_sql_server(
                                        f"UPDATE approve_file SET finish_flag=2,finish_status=N'已退回',popup_flag=1 "
                                        f"WHERE ID='{job_id}' ")
                                    count = count + res1 + res2
                                    update_sql_server(f"UPDATE approve_job SET job_flag=2 WHERE job_id={job_id} ")
                                else:
                                    sg.popup_auto_close('审批失败,审批人与当前用户不一致', title='审批失败', font=('宋体', 18))

                            if file_type_value == 2:
                                sql = f"UPDATE approve_job SET user_action=N'{user_action}',user_content=N'{multi_input}', " \
                                      f"enter_time='{enter_time}',finish_flag=1,job_flag=2 " \
                                      f"WHERE job_id={job_id} and id_number='{login_user}' "
                                res1 = update_sql_server(sql)
                                res2 = update_sql_server(
                                    f"UPDATE approve_file SET finish_flag=2,finish_status=N'已退回',popup_flag=1 "
                                    f"WHERE ID='{job_id}' ")
                                count = count + res1 + res2
                                update_sql_server(f"UPDATE approve_job SET job_flag=2 WHERE job_id={job_id} ")

                    if len(seleted_row) * 2 == count:
                        pop_times = 1  # 写入完成，开启任务弹屏
                        sg.popup_auto_close('操作成功', title='成功', font=('宋体', 18))
                    else:
                        sg.popup_auto_close('操作失败', title='失败', font=('宋体', 18))

                    # 更新待审批清单
                    # wait_approve_flag = 1
                    sql = f"SELECT {approve_file_para} FROM approve_file " \
                          f"WHERE ID IN (SELECT job_id FROM approve_job WHERE job_flag=1 and id_number='{login_user}') "
                    list_table_wait_approve = get_db_table(sql)
                    list_table = list_table_wait_approve
                    window1['TABLE2'].update(values=list_table)

        if event == '清空':
            window1['MT-OUTPUT'].update('')

        if check_job_flag == 1:
            query_job_count += 1
            # print('查询任务计时:', query_job_count)
            if query_job_count == 2:
                query_job_count = 0
                sql = f"SELECT job_id,job_flag FROM approve_job WHERE job_flag=1 and id_number='{login_user}'"
                res_flag = exec_sql_server_utf8_no_log(sql)
                try:
                    if len(res_flag) != 0:
                        window1['NEW-JOB-TEXT'].update('您有新的文件需要审批，请及时处理!')
                        if pop_times == 1:
                            sg.popup_ok(f'您有{len(res_flag)}份文件需要审批，请及时处理!', title='注意', font=('宋体', 18),
                                        keep_on_top=True)
                            pop_times = 0
                            wait_approve_flag = 1
                            # sg.SystemTray.notify(f'您有{len(res_flag)}份的文件需要审批，请及时处理!', '')
                    else:
                        window1['NEW-JOB-TEXT'].update('')
                        window1.refresh()
                except Exception as e:
                    sg.popup_error(e, title='错误', font=('宋体', 18))

        if check_popup_flag == 1:
            sql1 = f"SELECT file_name,finish_status FROM approve_file WHERE popup_flag=1 and id_number='{login_user}'"
            res_file_name = exec_sql_server_utf8_no_log(sql1)
            # print(res_file_name)
            try:
                if len(res_file_name) != 0:
                    for item in res_file_name:
                        filename = item[0]
                        finish_status = item[1]
                        if finish_status == '已退回':
                            sg.popup_ok(f'您的文件 {filename} 已退回，请及时处理!', title='文件已退回', font=('宋体', 18),
                                        keep_on_top=True, text_color='red')
                        else:
                            sg.popup_ok(f'您的文件 {filename} 已审批完成，请及时打印!', title='审批已完成', font=('宋体', 18),
                                        keep_on_top=True)
                        update_sql_server(f"UPDATE approve_file SET popup_flag=0 "
                                          f"WHERE file_name LIKE '%{filename}%' and id_number='{login_user}' ")
            except Exception as e:
                sg.popup_error(e, title='错误', font=('宋体', 18))

        if event == 'FRAME5_QUERY' or frame5_query_flag == 1:
            frame5_query_flag = 0
            lj_code = window1['FILE_CODE_LJ'].get()
            lj_pro_code = window1['PRO_CODE_LJ'].get()
            lj_rwl = window1['RWL_LJ'].get()
            lj_remark = window1['REMARK_LJ'].get()
            lj_type = window1['TYPE_LJ'].get()
            lj_time = window1['TIME_LJ'].get()
            table_para = 'ID,file_code,product_code,rwl,createdate,validdate,station,lj_type,remark'
            sql = f"SELECT {table_para} FROM lj_document WHERE file_code LIKE '%{lj_code}%' and product_code LIKE '%{lj_pro_code}%' " \
                  f"and rwl LIKE '%{lj_rwl}%' and lj_type LIKE '%{lj_type}%' and remark LIKE '%{lj_remark}%' " \
                  f"and createdate >= '{lj_time}' ORDER BY ID"
            list_table_lj = exec_sql_server_utf8(sql)
            window1['TABLE-LJ'].update(values=list_table_lj)

        if event == 'TABLE-LJDouble':
            seleted_row = values['TABLE-LJ']

            if len(seleted_row) == 0:
                sg.popup_auto_close('行不能为空,请选择对应行', title='注意', font=('宋体', 18))
            else:
                if sec_window is not None:
                    sec_window.close()
                window_lj_info = mw.make_window_lj_info()
                sec_window = window_lj_info
                if len(list_table_lj) > 0:
                    row_values = list_table_lj[seleted_row[0]]
                    id = row_values[0]
                    sql = f"SELECT * FROM lj_document WHERE ID={id} "
                    res = exec_sql_server_utf8(sql)
                    lj_id = res[0][0]
                    lj_code = res[0][1]
                    lj_pro_code = res[0][2]
                    lj_rwl = res[0][3]
                    lj_start_time = res[0][4]
                    lj_stop_time = res[0][5]
                    lj_info = res[0][6]
                    lj_station = res[0][7]
                    lj_type = res[0][8]
                    lj_remark = res[0][9]
                    window_lj_info['LJ-ID'].update(lj_id)
                    window_lj_info['LJ-CODE'].update(lj_code)
                    window_lj_info['LJ-PRO-CODE'].update(lj_pro_code)
                    window_lj_info['LJ-RWL'].update(lj_rwl)
                    window_lj_info['LJ-START-TIME'].update(lj_start_time)
                    window_lj_info['LJ-STOP-TIME'].update(lj_stop_time)
                    window_lj_info['LJ-STATION'].update(lj_station)
                    window_lj_info['LJ-TYPE'].update(lj_type)
                    window_lj_info['LJ-REMARK'].update(lj_remark)
                    window_lj_info['LJ-INFOMATION'].update(lj_info)

        if event == 'LJ-INFO-SAVE':
            lj_id = values['LJ-ID']
            lj_code = values['LJ-CODE']
            lj_pro_code = values['LJ-PRO-CODE']
            lj_rwl = values['LJ-RWL']
            lj_start_time = values['LJ-START-TIME']
            lj_stop_time = values['LJ-STOP-TIME']
            lj_info = values['LJ-INFOMATION'].strip()
            lj_station = values['LJ-STATION']
            lj_type = values['LJ-TYPE']
            lj_remark = values['LJ-REMARK']
            sql = f"UPDATE lj_document SET file_code='{lj_code}',product_code='{lj_pro_code}',rwl='{lj_rwl}'," \
                  f"createdate='{lj_start_time}',validdate='{lj_stop_time}',lj_info=N'{lj_info}',station='{lj_station}'," \
                  f"lj_type='{lj_type}',remark='{lj_remark}' WHERE ID={lj_id}  "
            res = update_sql_server(sql)
            if res == 1:
                window_lj_info.close()
                frame5_query_flag = 1
                sg.popup_auto_close('保存成功', title='成功', font=('宋体', 18))
            else:
                sg.popup_auto_close('保存失败', title='失败', font=('宋体', 18))

        if event == '打开临技-JZ':
            lj_path = r'\\10.10.80.80\share\宇视\临技\已转换'
            seleted_row = values['TABLE-LJ']

            if len(seleted_row) == 0:
                sg.popup_auto_close('行不能为空,请选择对应行', title='注意', font=('宋体', 18))
            else:
                row_values = list_table_lj[seleted_row[0]]
                id = row_values[0]
                sql = f"SELECT file_code FROM lj_document WHERE ID={id} "
                res = exec_sql_server_utf8(sql)
                if len(res) != 0:
                    lj_file_path = lj_path + '\\' + res[0][0] + '.xlsx'
                    print(lj_file_path)
                    try:
                        os.startfile(lj_file_path)
                    except Exception as e:
                        sg.popup_error(e, title='错误', font=('宋体', 18))

        if event == '打开临技':
            lj_path = r'\\10.10.80.80\share\宇视\临技\已处理'
            seleted_row = values['TABLE-LJ']

            if len(seleted_row) == 0:
                sg.popup_auto_close('行不能为空,请选择对应行', title='注意', font=('宋体', 18))
            else:
                row_values = list_table_lj[seleted_row[0]]
                id = row_values[0]
                sql = f"SELECT file_code FROM lj_document WHERE ID={id} "
                res = exec_sql_server_utf8(sql)
                if len(res) != 0:
                    lj_file_path = lj_path + '\\' + res[0][0] + '.xls'
                    print(lj_file_path)
                    try:
                        os.startfile(lj_file_path)
                    except Exception as e:
                        sg.popup_error(e, title='错误', font=('宋体', 18))

        if event == 'LJ-CONVER':
            lj_path = r'\\10.10.80.80\share\宇视\临技\已处理'
            lj_conver_path = r'\\10.10.80.80\share\宇视\临技\已转换'
            seleted_row = values['TABLE-LJ']

            if len(seleted_row) == 0:
                sg.popup_auto_close('行不能为空,请选择对应行', title='注意', font=('宋体', 18))
            else:
                row_values = list_table_lj[seleted_row[0]]
                id = row_values[0]
                sql = f"SELECT file_code FROM lj_document WHERE ID={id} "
                res = exec_sql_server_utf8(sql)
                if len(res) != 0:
                    lj_file_path = lj_path + '\\' + res[0][0] + '.xls'
                    print(lj_file_path)
                    ys_conver_jz_excel(lj_file_path, lj_conver_path)

        if event == 'DEL-LJ-INFO':
            seleted_row = values['TABLE-LJ']
            if len(seleted_row) == 0:
                sg.popup_auto_close('行不能为空,请选择对应行', title='注意', font=('宋体', 18))
            else:
                print(seleted_row)
                for num in seleted_row:
                    row_values = list_table_lj[num]
                    id = row_values[0]
                    sql = f"DELETE lj_document WHERE ID={id} "
                    res = update_sql_server(sql)
                sg.popup_auto_close('删除成功', title='成功', font=('宋体', 18))
                frame5_query_flag = 1

        if event == '临技审批详情':
            seleted_row = values['TABLE-LJ']

            if len(seleted_row) == 0:
                sg.popup_auto_close('行不能为空,请选择对应行', title='注意', font=('宋体', 18))
            else:
                if sec_window is not None:
                    sec_window.close()
                window_approve_info = mw.make_winodw_approve_info()
                sec_window = window_approve_info
                window_approve_info['TABLE-APP-INFO'].bind("<Double-Button-1>", 'Double')
                if login_user == 'root':
                    window_approve_info['APPROVE-RETRACT'].update(visible=True)

                if len(list_table_lj) > 0:
                    row_values = list_table_lj[seleted_row[0]]
                    id = row_values[0]
                    file_code = row_values[1]
                    res = exec_sql_server_utf8(f"SELECT job_id FROM lj_document WHERE ID={id} ")
                    if len(res) != 0:
                        job_id = res[0][0]
                        print('任务ID:', job_id)
                        if job_id is not None:
                            uname = exec_sql_server_utf8(f"SELECT Uname FROM approve_file WHERE ID='{job_id}' ")
                            sql = f"SELECT b.name,a.user_action,a.user_content,a.enter_time " \
                                  f"FROM approve_job AS a JOIN member AS b ON a.id_number = b.id_number " \
                                  f"where a.job_id={job_id} ORDER BY approve_level"
                            list_table3 = get_db_table(sql)
                            window_approve_info['APPROVE-TEXT-INFO'].update(file_code)
                            window_approve_info['TABLE-APP-INFO'].update(values=list_table3)
                            if len(uname) != 0:
                                window_approve_info['APPROVE-NAME-INFO'].update(uname[0][0])

        if event == 'APPROVE-RETRACT':
            s_row = values['TABLE-APP-INFO']
            print('任务ID:', job_id)
            if len(s_row) > 0:
                row_num = s_row[0]
                name_cn = list_table3[row_num][0]
                name_id = exec_sql_server_utf8(f"select id_number from member WHERE name=N'{name_cn}' ")
                id_number = name_id[0][0]
                print(id_number)
                sql = f"update approve_job set user_action='',user_content='',enter_time='',finish_flag=0,job_flag=1 " \
                      f"where job_id={job_id} and id_number='{id_number}' "
                sql1 = f"update approve_file set finish_flag=0, finish_status=N'待{name_cn}审批' where id={job_id} "
                res1 = update_sql_server(sql)
                res2 = update_sql_server(sql1)
                if res1 == 1 and res2 == 1:
                    sg.popup_auto_close('撤回成功', title='成功', font=('宋体', 18))
                else:
                    sg.popup_auto_close('撤回失败', title='失败', font=('宋体', 18))

        if event == '日期':
            get_time = sg.popup_get_date(close_when_chosen=True, location=(320, 220))
            time_lj = str(get_time[2]) + '-' + str(get_time[0]) + '-' + str(get_time[1])
            window1['TIME_LJ'].update(time_lj)

        if event == 'DATE-BUTTON':
            get_time = sg.popup_get_date(close_when_chosen=True, location=(1050, 230))
            time_lj = str(get_time[2]) + '-' + str(get_time[0]) + '-' + str(get_time[1])
            window1['DATE-INPUT'].update(time_lj)

        if event == 'LOOK-MES-LJTABLE':
            if window_mes_ljtable is not None:
                window_mes_ljtable.close()
            window_mes_ljtable = mw.make_winodw_mes_ljtable()
            # sec_window = window_mes_ljtable
            sql = f"SELECT ID,Rwl,ItemCode1,beizhu FROM LJSHTable "
            list_ljtable = get_db_table_mes(sql)
            window_mes_ljtable['TABLE-MES'].update(values=list_ljtable)
            window_mes_ljtable['TABLE-MES'].bind("<Double-Button-1>", 'Double')
            window_mes_ljtable['MES-RWL-INPUT'].bind("<Return>", 'Return')
            window_mes_ljtable['MES-CODE-INPUT'].bind("<Return>", 'Return')
            window_mes_ljtable['MES-REMARK-INPUT'].bind("<Return>", 'Return')

        if event == 'TABLE-MESDouble' or event == 'MES-EDIT':
            select_row = values['TABLE-MES']
            if len(select_row) == 0:
                sg.popup_auto_close('请选择需要修改的行', title='注意', font=('宋体', 18))
            else:
                if sec_window is not None:
                    sec_window.close()
                window_mes_edit = mw.make_window_mes_edit()
                sec_window = window_mes_edit
                row_number = select_row[0]
                row_value = list_ljtable[row_number]
                id_value = row_value[0]
                rwl = row_value[1]
                code = row_value[2]
                remark = row_value[3]
                res = exec_sql_server_mes(f"SELECT SCName,SCTime FROM LJSHTable WHERE ID={id_value} ")
                sc_user = res[0][0]
                sc_time = res[0][1]
                window_mes_edit['MES-EDIT-ID'].update(id_value)
                window_mes_edit['MES-EDIT-RWL'].update(rwl)
                window_mes_edit['MES-EDIT-PRO-CODE'].update(code)
                window_mes_edit['MES-EDIT-USER'].update(sc_user)
                window_mes_edit['MES-EDIT-TIME'].update(sc_time)
                window_mes_edit['MES-EDIT-REMARK'].update(remark)

        if event == 'MES-EDIT-SAVE':
            id = values['MES-EDIT-ID']
            rwl = values['MES-EDIT-RWL']
            proc_code = values['MES-EDIT-PRO-CODE']
            sc_user = values['MES-EDIT-USER']
            sc_time = values['MES-EDIT-TIME']
            remark = values['MES-EDIT-REMARK'].strip()
            sql = f"UPDATE LJSHTable SET Rwl='{rwl}',ItemCode1='{proc_code}',beizhu=N'{remark}',SCName='{sc_user}'," \
                  f"SCTime='{sc_time}' WHERE ID={id} "
            res = update_sql_server_mes(sql)
            if res == 1:
                window_mes_edit.close()
                ljtable_query_flag = 1
                sg.popup_auto_close('保存成功', title='成功', font=('宋体', 18))
            else:
                sg.popup_auto_close('保存失败', title='失败', font=('宋体', 18))

        if event == 'MES-ADD':
            if sec_window is not None:
                sec_window.close()
            window_mes_add = mw.make_window_mes_add()
            sec_window = window_mes_add

        if event == 'MES-ADD-SAVE':
            count = 0
            number = 0
            res = 0
            mes_rwls = values['MES-RWL']
            mes_pro_code = values['MES-PRO-CODE']
            mes_remark = values['MES-REMARK'].strip()
            sc_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
            insert_para = '(Rwl, ItemCode1, beizhu, SCName, SCTime)'
            if mes_pro_code == '/' and mes_rwls != '':
                mes_rwl = mes_rwls.split('/')
                number = len(mes_rwl)
                for rwl in mes_rwl:
                    insert_value = (rwl, mes_pro_code, mes_remark, login_user, sc_time)
                    res = update_sql_server_mes(f"INSERT INTO LJSHTable {insert_para} VALUES {insert_value}")
                    count += 1
                if count == number:
                    window_mes_add.close()
                    ljtable_query_flag = 1
                    sg.popup_auto_close('数据写入成功', title='成功', font=('宋体', 18))
                else:
                    sg.popup_auto_close('数据写入失败', title='失败', font=('宋体', 18))
            elif mes_pro_code != '' and mes_rwls == '/':
                pro_code = mes_pro_code.split('/')
                number = len(pro_code)
                for code in pro_code:
                    ys_code = 'K-YS-' + code
                    insert_value = (mes_rwls, ys_code, mes_remark, login_user, sc_time)
                    res = update_sql_server_mes(f"INSERT INTO LJSHTable {insert_para} VALUES {insert_value}")
                    count += 1

                if count == number:
                    window_mes_add.close()
                    ljtable_query_flag = 1
                    sg.popup_auto_close('数据写入成功', title='成功', font=('宋体', 18))
                else:
                    sg.popup_auto_close('数据写入失败', title='失败', font=('宋体', 18))

            else:
                sg.popup_error('任务令或产品编码格式有误，请检查', title='格式错误', font=('宋体', 18))

        if event == 'MES-DEL':
            select_row = values['TABLE-MES']
            if len(select_row) == 0:
                sg.popup_auto_close('请选择需要删除的行', title='注意', font=('宋体', 18))
            else:
                row_number = select_row[0]
                row_value = list_ljtable[row_number]
                id_value = row_value[0]
                res = update_sql_server_mes(f"DELETE LJSHTable WHERE ID={id_value} ")
                if res == 1:
                    ljtable_query_flag = 1
                    sg.popup_auto_close('删除成功', title='成功', font=('宋体', 18))
                else:
                    sg.popup_auto_close('删除失败', title='失败', font=('宋体', 18))

        if event == 'MES-QUERY' or ljtable_query_flag == 1:
            ljtable_query_flag = 0
            rwl_para = window_mes_ljtable['MES-RWL-INPUT'].get()
            pro_code_para = window_mes_ljtable['MES-CODE-INPUT'].get()
            remark_para = window_mes_ljtable['MES-REMARK-INPUT'].get()
            sql = f"SELECT ID,Rwl,ItemCode1,beizhu FROM LJSHTable " \
                  f"WHERE Rwl LIKE '%{rwl_para}%' and ItemCode1 LIKE '%{pro_code_para}%' and beizhu LIKE '%{remark_para}%' "
            list_ljtable = get_db_table_mes(sql)
            window_mes_ljtable['TABLE-MES'].update(values=list_ljtable)

        if event == 'LJ-FILE-OUTPUT':
            print(list_table_lj)
            now_time = time.strftime('%Y-%m-%d-%H-%M-%S', time.localtime(time.time()))
            new_lj_excel = excel_path + '\\' + 'LJ' + f'-{now_time}.xlsx'
            wb = Workbook()  # 创建工作薄对象
            ws = wb.create_sheet('临技', 0)  # 创建工作表对象,0对应第1个sheet，1为第2个sheet
            if len(list_table_lj) != 0:
                try:
                    for i in range(len(list_table_lj)):
                        item = list_table_lj[i]
                        print(item)
                        write_row_by_list(ws, i + 1, item)
                    wb.save(new_lj_excel)
                    logger.info(f'导出成功,路径:{new_lj_excel}')
                    sg.popup_ok(f'路径:{new_lj_excel}', title='导出成功', font=('宋体', 18), )

                except Exception as result:
                    sg.popup_error(result, font=('宋体', 18))
            else:
                sg.popup_auto_close('请先查询，再导出', title='注意', font=('宋体', 18))

    window1.close()
