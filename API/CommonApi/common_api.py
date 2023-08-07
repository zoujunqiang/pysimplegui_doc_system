# encoding:utf-8
"""
@file = common_api
@author = zouju
@create_time = 2023-08-07- 14:22
"""
import datetime
import hashlib
import os
import shutil
import subprocess
from datetime import time

import PySimpleGUI as sg
import qrcode
import win32com.client as win32
from openpyxl import Workbook
from openpyxl import load_workbook

from API.SqlExec.sql_exec import exec_sql_server_utf8, update_sql_server, exec_sql_server_mes250
from CONFIG.settings import server_path, lj_path, jz_file_path, ys_xlsx_file_path
from API.Log.log import logger


def myPopen(cmd):
    subprocess.Popen(cmd, shell=True, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

    # return proc.stdout.read().decode()


def upload_to_server(old_filepath, newfile):
    global id, file_code, file_name, version
    file_para2 = ('LS',)
    file_para4 = ('JZ-P', 'YSWL', 'BLWL', 'FHWL', 'YDWL', 'L202')
    file_para6 = ('JZ-EWI',)
    file_para8 = (
        'JZ-WI-HR', 'JZ-WI-PD', 'JZ-WI-PL', 'JZ-WI-PM', 'JZ-WI-PR', 'JZ-WI-PU ', 'JZ-WI-QC', 'JZ-WI-RD',
        'JZ-WI-SA', 'JZ-WI-WX', 'JZ-WI-XZ', 'JZ-WI-FN', 'JZ-WI-ME', 'JZ-WI-PE', 'JZ-WI-TE', 'JZ-WI-IE',
        'JZ-QR-FN', 'JZ-QR-PU', 'JZ-QR-PR', 'JZ-QR-LW', 'JZ-ER-LW', 'JZ-QR-SM', 'JZ-QR-GM', 'JZ-QR-XZ',
        'JZ-QR-PL', 'JZ-QR-QC', 'JZ-ER-QC', 'JZ-QR-HR', 'JZ-QR-SA', 'JZ-QR-PM', 'JZ-QR-RD', 'JZ-QR-IE',
        'JZ-QR-ME', 'JZ-QR-PE', 'JZ-QR-TE', 'JZ-QR-PD', 'JZ-WI-GM')
    file_para9 = ('JZ-WI-DIP', 'JZ-WI-SMT', 'JZ-QR-SMT', 'JZ-QR-DIP', 'JZ-QR-DCC')
    dict_file_code = {
        'LS': '\临时文件', 'JZ-P': '\二阶文件(新)', 'YSWL': '\外来文件\宇视文件', 'BLWL': '\外来文件\博联文件',
        'FHWL': '\外来文件\烽火文件', 'YDWL': '\外来文件\枪机移动', 'JZ-EWI': '\三阶文件\EWI(环境，质量)',
        'JZ-WI-HR': '\三阶文件\HR(人事)', 'JZ-WI-PD': '\三阶文件\生产\PD', 'JZ-WI-PL': '\三阶文件\PL(计划)',
        'JZ-WI-PM': '\三阶文件\PM(业务)', 'JZ-WI-PR': '\三阶文件\PR(项目)', 'JZ-WI-PU': '\三阶文件\PU(采购)',
        'JZ-WI-QC': '\三阶文件\QC(品质)', 'JZ-WI-RD': '\三阶文件\RD(研发)', 'JZ-WI-SA': '\三阶文件\SA(市场)',
        'JZ-WI-WX': '\三阶文件\WX(维修)', 'JZ-WI-XZ': '\三阶文件\XZ(行政)', 'JZ-WI-FN': '\三阶文件\FN(财务)',
        'JZ-WI-ME': '\三阶文件\工程部\ME', 'JZ-WI-PE': '\三阶文件\工程部\PE', 'JZ-WI-TE': '\三阶文件\工程部\TE',
        'JZ-WI-IE': '\三阶文件\工程部\IE', 'JZ-QR-FN': '\四阶文件\财务部FN', 'JZ-QR-PU': '\四阶文件\采购PU',
        'JZ-QR-PR': '\四阶文件\产品部PR', 'JZ-QR-LW': '\四阶文件\法务部LW', 'JZ-ER-LW': '\四阶文件\法务部LW',
        'JZ-QR-GM': '\四阶文件\管代GM', 'JZ-QR-XZ': '\四阶文件\行政部XZ', 'JZ-QR-QC': '\四阶文件\品质部QR',
        'JZ-ER-QC': '\四阶文件\品质部QR', 'JZ-QR-HR': '\四阶文件\人事HR', 'JZ-QR-SA': '\四阶文件\市场部SA',
        'JZ-QR-PM': '\四阶文件\项目PM', 'JZ-QR-RD': '\四阶文件\研发部RD', 'JZ-QR-IE': '\四阶文件\工程部\IE',
        'JZ-QR-ME': '\四阶文件\工程部\ME', 'JZ-QR-PE': '\四阶文件\工程部\PE', 'JZ-QR-TE': '\四阶文件\工程部\TE',
        'JZ-QR-PD': '\四阶文件\生产部\PD', 'JZ-QR-PL': '\四阶文件\计划物流部PL', 'JZ-QR-SM': '\四阶文件\供管部SM',
        'JZ-WI-DIP': '\三阶文件\生产\DIP', 'JZ-WI-SMT': '\三阶文件\工程部\SMT', 'JZ-QR-DCC': '\四阶文件\DCC',
        'JZ-QR-SMT': '\四阶文件\生产部\SMT', 'JZ-QR-DIP': '\四阶文件\生产部\DIP', 'JZ-WI-GM': '\三阶文件\GM(管代)',
        '': ''
    }

    file_dict1 = {'level2_document': '二阶文件', 'level3_document': '三阶文件', 'level4_document': '四阶文件',
                  'foreign_document': '外来文件', 'temp_document': '临时文件', 'document_issue': '文件发放',
                  'lj_document': '临技文件', '': ''}

    if newfile[0:2] in file_para2:
        new_path = server_path + dict_file_code[f'{newfile[0:2]}'] + '\\' + newfile
        res = copyfile(old_filepath, new_path)
        if res == 1:
            return new_path
        else:
            return 0

    elif newfile[0:4] in file_para4:
        if newfile[0:4] == 'L202' or newfile[0:4] == 'C202':
            # table_name = 'lj_document'
            new_path = lj_path + '\\' + newfile

        elif newfile[0:4] == 'JZ-P':
            # table_name = 'level2_document'
            new_path = server_path + dict_file_code[f'{newfile[0:4]}'] + '\\' + newfile

        else:
            # table_name = 'foreign_document'
            new_path = server_path + dict_file_code[f'{newfile[0:4]}'] + '\\' + newfile

        res = copyfile(old_filepath, new_path)

        if res == 1:
            return new_path
        else:
            return 0

    elif newfile[0:6] in file_para6:
        # table_name = 'level3_document'
        new_path = server_path + dict_file_code[f'{newfile[0:6]}'] + '\\' + newfile
        res = copyfile(old_filepath, new_path)
        if res == 1:
            return new_path
        else:
            return 0

    elif newfile[0:8] in file_para8:
        new_path = server_path + dict_file_code[f'{newfile[0:8]}'] + '\\' + newfile
        res = copyfile(old_filepath, new_path)
        if res == 1:
            return new_path
        else:
            return 0

    elif newfile[0:9] in file_para9:

        new_path = server_path + dict_file_code[f'{newfile[0:9]}'] + '\\' + newfile
        res = copyfile(old_filepath, new_path)
        if res == 1:
            return new_path
        else:
            return 0

    elif '0302' in newfile:
        new_path = r'\\10.10.80.80\share\宇视\原型机\原型机试产报告\auto_copy'
        res = copyfile(old_filepath, new_path)
        if res == 1:
            return new_path
        else:
            return 0

    else:
        sg.popup_auto_close('文件名不符合命名规则,请联系管理员！', title='文件名不符合命名规则', font=('宋体', 18))
        logger.info('文件名不符合命名规则,请联系管理员！')
        return 0


def write_db_from_file(new_file):
    db_tablename = ['level2_document', 'level3_document', 'level4_document', 'temp_document']
    count = 0
    file_splited = new_file.split('_')
    if len(file_splited) > 2:
        file_code = file_splited[0]
        file_version = file_splited[1]
        file_name = file_splited[2]
        file_date = time.strftime('%Y-%m-%d', time.localtime(time.time()))
        for table in db_tablename:
            sql = f"SELECT ID FROM {table} WHERE file_code='{file_code}' "
            res = exec_sql_server_utf8(sql)
            if len(res) > 0:
                count = 1
                id = res[0][0]
                sql1 = f"UPDATE {table} set file_name=N'{file_name}',version='{file_version}',createdate='{file_date}' " \
                       f"WHERE ID={id}"
                res = update_sql_server(sql1)
                if res == 1:
                    sg.popup_auto_close('更新数据成功', title='失败', font=('宋体', 18))
                else:
                    sg.popup_auto_close('更新数据失败', title='失败', font=('宋体', 18))
                    print('更新数据失败')
        if count == 0:
            sg.popup_auto_close(f'未查询到文件编码:{file_code}', title='失败', font=('宋体', 18))

    elif '0302' in new_file:
        pass

    else:
        sg.popup_auto_close(f'{new_file}文件格式有误,写入数据失败', title='失败', font=('宋体', 18))


def copyfile(path, new_filepath):
    try:
        shutil.copy(path, new_filepath)
        logger.info(f'文件临时地址:{path}')
        logger.info(f'{new_filepath} 归档成功')
        return 1

    except Exception as result:
        logger.info(f'{new_filepath} 归档失败:{result}')
        sg.popup_error(result, font=('宋体', 18))
        return 0


def get_FileSize(filePath):
    try:
        fsize = os.path.getsize(filePath)
        kb = 1024
        fsize = fsize / kb
        return round(fsize, 2)
    except Exception as e:
        sg.popup_error(e, font=('宋体', 18))


def computerMD5(message):
    m = hashlib.md5()
    m.update(message.encode(encoding='utf-8'))
    return m.hexdigest()


def write_row_by_list(ws, row, list_value):  # 把list_value中的数据依次写入第row行，写入第一张表
    for i in range(0, len(list_value)):  # len(list)为列表的长度
        ws.cell(row, i + 1).value = list_value[i]  # 从第1列开始



def open_file(path, name):  # 查找文件
    try:
        for item in os.listdir(path):
            item_path = os.path.join(path, item)
            if os.path.isdir(item_path):
                open_file(item_path, name)
            elif os.path.isfile(item_path):
                if name in item:
                    os.startfile(item_path)
                    break

                    # logger.info(f"打开文件:{item_path}")
                    # return 1
    except Exception as results:
        sg.popup_error(results, font=('宋体', 18))
        return 0


def make_qr_code(content, save_path=None):
    qr_code_maker = qrcode.QRCode(version=5,
                                  error_correction=qrcode.constants.ERROR_CORRECT_M,
                                  box_size=8,
                                  border=4,
                                  )
    qr_code_maker.add_data(data=content)
    qr_code_maker.make(fit=True)
    img = qr_code_maker.make_image(fill_color="black", back_color="white")
    if save_path:
        img.save(save_path)
    else:
        img.show()  # 中间图不显示


def list_col_value(sheet, column, lenth):
    list_col_data = []  # 定义列表，存储excel 对应列的值
    for i in range(1, lenth + 1):
        cell_value = sheet.cell(i, column).value
        list_col_data.append(cell_value)
    return list_col_data



def ys_conver_jz_excel(ys_xls_file_path, new_lj_path):
    global excel
    xls_open_flag = 0
    pro_code = []
    num_list = []
    num1_list = []
    now_ver = []
    after_ver = []
    mat_code = []
    mat_des = []
    original_num = []
    now_num = []
    delete_position = []
    add_possition = []
    remark_info = []

    lj_name = ys_xls_file_path.split('\\')
    new_lj_name = new_lj_path + '\\' + lj_name[-1] + "x"
    shutil.copyfile(jz_file_path, new_lj_name)

    # xls转xlsx
    ys_file_path = ys_xlsx_file_path + '\\' + lj_name[-1] + "x"

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    try:
        wb = excel.Workbooks.Open(ys_xls_file_path)
        wb.SaveAs(ys_file_path, FileFormat=51)  # FileFormat = 51 is for .xlsx extension
        wb.Close()  # FileFormat = 56 is for .xls extension
        excel.Application.Quit()
    except Exception as e:
        xls_open_flag = 1
        sg.popup_error(e, title='错误', font=('宋体', 18))

    if xls_open_flag == 0:
        # jz obj
        wb_jz = load_workbook(jz_file_path)
        sheets_jz = wb_jz.sheetnames
        sheet_jz = wb_jz['临技']
        # ys obj
        wb_ys = load_workbook(ys_file_path)
        sheets_ys = wb_ys.sheetnames
        sheet_ys = wb_ys[sheets_ys[0]]
        mrow_ys = sheet_ys.max_row  # 获取行数
        mcol_ys = sheet_ys.max_column  # 获取列数
        col1_list = list_col_value(sheet_ys, 1, mrow_ys)
        col2_list = list_col_value(sheet_ys, 2, mrow_ys)
        num_code = col1_list.index('申请单号')
        num_rwl = col1_list.index('加工任务令号1（非改制类最多添加10行）')
        num_type = col1_list.index('类别')
        num_date = col1_list.index('请指定“指定时间内有效的临技”的有效截止时间')

        for i in range(mcol_ys):
            for j in range(mrow_ys):
                cell_value = sheet_ys.cell(j + 1, i + 1).value
                if isinstance(cell_value, str):  # 判断cell_value是否为str类型
                    # print(cell_value)
                    if cell_value == '更改前项目编码':
                        code_col = i + 1
                        code_row = j + 1
                        for k in range(mcol_ys):
                            code_cell = sheet_ys.cell(code_row + k + 1, code_col).value
                            # if isinstance(code_cell, str):
                            if code_cell is not None:
                                row_number = code_row + k + 1
                                pro_code.append(code_cell)
                                num_list.append(row_number)
                            else:
                                break
                    if cell_value == '更改前编码版本':
                        code_col = i + 1
                        for k in range(len(num_list)):
                            ver_cell = sheet_ys.cell(num_list[k], code_col).value
                            now_ver.append(ver_cell)
                    if cell_value == '更改后编码版本':
                        code_col = i + 1
                        for k in range(len(num_list)):
                            ver_cell = sheet_ys.cell(num_list[k], code_col).value
                            after_ver.append(ver_cell)

        for i in range(mcol_ys):
            for j in range(mrow_ys):
                cell_value = sheet_ys.cell(j + 1, i + 1).value
                if isinstance(cell_value, str):  # 判断cell_value是否为str类型
                    # print(cell_value)
                    if cell_value == '项目编码':
                        code_col = i + 1
                        code_row = j + 1
                        for k in range(mcol_ys):
                            code_cell = sheet_ys.cell(code_row + k + 1, code_col).value
                            if code_cell is not None:
                                row_number = code_row + k + 1
                                mat_code.append(code_cell)
                                num1_list.append(row_number)
                            else:
                                break
                    if cell_value == '项目描述':
                        code_col = i + 1
                        for k in range(len(num1_list)):
                            ver_cell = sheet_ys.cell(num1_list[k], code_col).value
                            mat_des.append(ver_cell)

                    if cell_value == '原数量':
                        code_col = i + 1
                        for k in range(len(num1_list)):
                            ver_cell = sheet_ys.cell(num1_list[k], code_col).value
                            original_num.append(ver_cell)

                    if cell_value == '现数量':
                        code_col = i + 1
                        for k in range(len(num1_list)):
                            ver_cell = sheet_ys.cell(num1_list[k], code_col).value
                            now_num.append(ver_cell)

                    if cell_value == '删除位号':
                        code_col = i + 1
                        for k in range(len(num1_list)):
                            ver_cell = sheet_ys.cell(num1_list[k], code_col).value
                            delete_position.append(ver_cell)

                    if cell_value == '增加位号':
                        code_col = i + 1
                        for k in range(len(num1_list)):
                            ver_cell = sheet_ys.cell(num1_list[k], code_col).value
                            add_possition.append(ver_cell)

        for k in range(len(num1_list)):
            ver_cell = sheet_ys.cell(num1_list[k], 12).value
            remark_info.append(ver_cell)

        # 写入数据
        now_time = time.strftime('%Y-%m-%d', time.localtime(time.time()))
        sheet_jz['G3'] = now_time
        sheet_jz['I3'] = col2_list[num_code]
        sheet_jz['B4'] = '烧录 条码打印 SMT DIP 测试 包装'
        sheet_jz['G4'] = '物料 烧录软件 辅材 FT程式 工艺 其他'
        if col2_list[num_rwl] is None:
            write_date = time.strftime('%Y/%m/%d--', time.localtime(time.time()))
            valid_date = datetime.datetime.strptime(str(col2_list[num_date]), '%Y-%m-%d %H:%M:%S').strftime('%Y/%m/%d')
            sheet_jz['C70'] = '期间有效'
            sheet_jz['H70'] = str(write_date) + str(valid_date)
        else:
            sheet_jz['C70'] = '一次性有效'
            sheet_jz['H70'] = col2_list[num_rwl]

        for i in range(len(pro_code)):
            ys_pro_code = 'K-YS-' + pro_code[i]
            ys_pro_des = exec_sql_server_mes250(f"SELECT ItemName FROM basal_item where ItemCode='{ys_pro_code}' ")
            if len(ys_pro_des) == 0:
                ys_pro_des_value = ''
            else:
                ys_pro_des_value = ys_pro_des[0][0]
            sheet_jz.cell(7 + i, 2).value = pro_code[i]
            sheet_jz.cell(7 + i, 3).value = ys_pro_des_value
            sheet_jz.cell(7 + i, 7).value = ys_pro_code
            sheet_jz.cell(7 + i, 9).value = now_ver[i]
            sheet_jz.cell(7 + i, 10).value = after_ver[i]

        for i in range(len(mat_code)):
            sheet_jz.cell(39 + i, 2).value = mat_code[i]
            sheet_jz.cell(39 + i, 3).value = mat_des[i]
            sheet_jz.cell(39 + i, 6).value = original_num[i]
            sheet_jz.cell(39 + i, 7).value = now_num[i]
            sheet_jz.cell(39 + i, 8).value = delete_position[i]
            sheet_jz.cell(39 + i, 9).value = add_possition[i]
            sheet_jz.cell(39 + i, 10).value = remark_info[i]

        try:
            wb_jz.save(new_lj_name)
            print(f'已转换完成,新文件路径:{new_lj_name}')
            sg.popup_ok(f'已转换完成,新文件路径:{new_lj_name}', title='成功', font=('宋体', 18))

        except Exception as e:
            sg.popup_error(e, title='错误', font=('宋体', 18))